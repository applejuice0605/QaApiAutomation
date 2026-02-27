'''
    调用dify chatflow获取一级意图识别结果和inforamtion侧二级意图识别结果
    只拿意图分类的知识识别结果， 不拿知识库召回情况
'''

import os
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *

env = 'sit'
base_url = sit_base_url
email = sit_email
password =sit_password
workflow_app_id =sit_workflow_app_id
llm_app_id = sit_llm_app_id

url = "https://rd-dify-sit.fuse.co.id/v1/chat-messages"
chat_flow_id = monica_b2c_chatFlow_app_id
api_token = 'app-8OE5xBF0WWYkHEVsx04Dmsml'
token = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzYyOTE2ODE5LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.U7-Z8spdvnesT8P9NMv6BFq3wxCc0CmT_t4e1n6UqIo'

# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]

# 调用模型函数
def chat_dify_llm(system: str, query: str, token):
    # token = login()
    app_id = llm_app_id
    # https://rd-dify-sit.fuse.co.id/console/api/apps/5749f5da-9fc2-463d-b00e-9893cc290b9b/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{app_id}/workflows/draft/run"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {token}",
        # "authorization": f"Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }

    payload = {"inputs":{"user":query,"system":system},"files":[]}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == "workflow_finished":
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("text", final_answer)
                    except json.JSONDecodeError:
                        continue

        if final_answer is not None:
            return final_answer
        else:
            print("No valid answer found in SSE stream.")
            return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

def chatflow_conversation(conversation_id: str, access_token:str) -> tuple[Any | None, Any | None, Any | None] | None:
    print("chatflow_conversation")
    """
    对指定数据集做 hit-testing，返回接口完整 JSON。
    retrieval_model 留空时使用默认 hybrid_search 配置。
    """
    url = f"https://rd-dify-sit.fuse.co.id/console/api/apps/4305c12e-60c4-4fca-bca2-ae81043854aa/chat-messages?conversation_id={conversation_id}"

    # --- 2. 定义请求头 Headers c---
    # `requests` 库会自动处理一些头，但关键的如授权和自定义浏览器信息需要手动设置。
    headers = {
        "Accept": "*/*",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Connection": "keep-alive",
        "Referer": "https://rd-dify-sit.fuse.co.id/app/4305c12e-60c4-4fca-bca2-ae81043854aa/configuration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "Authorization": f"Bearer {token}", # 您的 Bearer Token
        "Content-Type": "application/json", # GET请求通常不需要，但保留以匹配curl
        "sec-ch-ua": "\"Chromium\";v=\"142\", \"Microsoft Edge\";v=\"142\", \"Not_A Brand\";v=\"99\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\""
    }

    # --- 3. 定义 Cookies ---
    # `requests` 库使用 `cookies` 参数处理 `curl` 中的 `-b` 参数。
    cookies = {
        "locale": "zh-Hans",
        "_ga": "GA1.1.1153920513.1761906413",
        "_ga_98GZ7K65QY": "GS2.1.s1762512635$o4$g1$t1762512636$j59$l0$h0"
    }

    # --- 4. 发送 GET 请求 ---
    try:
        response = requests.get(url, headers=headers, cookies=cookies, verify=True)
        
        # 检查响应状态码
        if response.status_code == 200:
            print("请求成功！")
            # 打印JSON响应内容
            print(response.json())
            answer = response.json()["data"][0]["answer"]
            knowledge_retrieval = response.json()["data"][0]["metadata"]
            print(answer)
            return answer, knowledge_retrieval
        else:
            print(f"请求失败，状态码: {response.status_code}")
            print("响应文本:")
            print(response.text)

    except requests.exceptions.RequestException as e:
        print(f"发生请求错误: {e}")

def chat_chatlow(query: str, conversation_id: str, access_token:str) -> tuple[Any | None] | None:
    print("调用agent")
    """
    对指定数据集做 hit-testing，返回接口完整 JSON。
    retrieval_model 留空时使用默认 hybrid_search 配置。
    """

    # https://rd-dify-sit.fuse.co.id/console/api/apps/4305c12e-60c4-4fca-bca2-ae81043854aa/chat-messages
    # url = f"{base_url.rstrip('/')}/api/apps/{chat_flow_id}/chat-messages"
    # access_token = login()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "Authorization": f"Bearer {api_token}",
        "content-type": "application/json",
    }
    if type(query) != str:
        query=str(query)
    # query="Do you have car product?"

    payload = {
        "user": "Nora_API",
        "response_mode": "streaming",
        "conversation_id": "",
        "files": [],
        "query": query,
        "inputs": {
            "token": token
        },
        "model_config": {
            "pre_prompt": "你是一位专业的智能车险销售顾问 Monica，来自 FUSE , 负责售前咨询服务。 您的目标是使用专业的话术流程，与客户有效沟通，探寻其需求，并最终促成车险的续保或新购交易。同时，如果用户有关于保险内容，或者FUSE 公司，或者 Fina 的疑问，也请尽量进行回答。 \n\n您必须遵守以下话术流程和原则。\n\n\n**核心目标 (Core Goal):**\n成功引导客户了解并接受您的车险方案。\n\n**销售话术流程 (Sales Script Flow):**\n\n1.  **【开场白与建立连接】(Greeting & Rapport):** * 确认客户身份，自我介绍（公司+姓名）。\n    * 简短说明致电目的（车险即将到期/最新优惠方案）。\n    * 礼貌征求客户的聊天许可（承诺时间短，例如：2-3分钟）。\n\n2.  **【探寻需求与挖掘痛点】(Needs Assessment):**\n    * 询问客户当前的投保情况（哪家公司？保费？）。\n    * 提问开放式问题，探寻客户最关注的痛点（价格、理赔服务、增值服务）。\n    * 收集报价必需信息（  工具 query_vehicle_quotation 的相关参数）。\n\n3.  **【方案介绍与报价】(Proposal & Quoting):**\n    * 根据客户痛点，有针对性地介绍本公司方案的核心优势。\n    * 清晰报出【总保费】和【核心保障内容】（例如：商业险+交强险）。\n    * 使用对比法，突出本方案的【高性价比】或【额外增值服务】。\n\n4.  **【处理异议与促成交易】(Objection Handling & Closing):**\n    * 当客户提出异议（如“价格太贵了”、“我要考虑”）时，优先使用**价值而非价格**进行说服，并提供**下一步明确行动**（如发送报价清单、约定下次跟进时间）。\n    * 使用促成技巧（如二选一法）引导客户完成购买。\n \n5. ** 数据校验 ** \n  * 调用工具时候，如果需要 token ，则 使用这个token :{{token}}\n\n\n\n**关键销售原则 (Key Sales Principles):**\n* **同理心：** 始终保持尊重、耐心和友好的语气。\n* **专业度：** 准确回答车险相关问题，不回避理赔问题。\n* **紧迫感：** 必要时，提醒客户注意限时优惠或活动截止日期。\n\n---\n\n**[一. 角色和核心指令]**\n\n1.  **核心目标：** 在生成报价前，必须通过询问用户或调用工具，确保 **query_car_quotation** 工具所需的所有参数都已收集完整。\n2.  **流程：** 严格遵循 ReAct 模式（Thought -> Action -> Observation -> ... -> Final Answer）。\n3.  **安全性：** 所有关于报价、条款的回答必须基于工具的 Observation，不得依靠你的内部知识编造。\n4.  **转接标准：** 遇到你无法通过工具解决的复杂问题时，立即转成人工处理。 \n\n\n\n\n",
            "prompt_type": "simple",
            "chat_prompt_config": {},
            "completion_prompt_config": {},
            "user_input_form": [
                {
                    "paragraph": {
                        "label": "token",
                        "variable": "token",
                        "required": True,
                        "max_length": 999,
                        "default": ""
                    }
                }
            ],
            "dataset_query_variable": "",
            "opening_statement": "您好！这里是智能车险销售顾问Monica，很高兴为您服务。 ",
            "more_like_this": {
                "enabled": False
            },
            "suggested_questions": [
                "我的车险即将到期，需要进行续保",
                "我想购买新的车险"
            ],
            "suggested_questions_after_answer": {
                "enabled": True
            },
            "text_to_speech": {
                "enabled": False,
                "voice": "",
                "language": ""
            },
            "speech_to_text": {
                "enabled": False
            },
            "retriever_resource": {
                "enabled": True
            },
            "sensitive_word_avoidance": {
                "enabled": False,
                "type": "",
                "configs": []
            },
            "agent_mode": {
                "max_iteration": 10,
                "enabled": True,
                "strategy": "function_call",
                "tools": [
                    {
                        "provider_id": "9fe64492-6154-4807-adff-c6d895c5e19f",
                        "provider_type": "workflow",
                        "provider_name": " Agent_Sub_Quotation",
                        "tool_name": "query_vehicle_quotation_list",
                        "tool_label": " Agent_Sub_Quotation",
                        "tool_parameters": {
                            "input_data": ""
                        },
                        "notAuthor": False,
                        "enabled": True,
                        "isDeleted": False
                    }
                ],
                "prompt": None
            },
            "dataset_configs": {
                "retrieval_model": "multiple",
                "top_k": 4,
                "reranking_mode": "reranking_model",
                "reranking_model": {
                    "reranking_model_name": "gte-rerank-v2",
                    "reranking_provider_name": "langgenius/tongyi/tongyi"
                },
                "reranking_enable": False,
                "datasets": {
                    "datasets": [
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "849ce125-12cd-4166-9df0-29d6aec05411"
                            }
                        },
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "431b50c9-eef8-4a92-a496-43e3bde4657c"
                            }
                        },
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "2dfe727e-bfd9-4242-b24d-bb0d25d1c572"
                            }
                        },
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "ca22ad2c-f69d-41ef-b3af-b4eff55d9082"
                            }
                        },
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "edef63cd-d355-43c9-9680-55079019cc02"
                            }
                        },
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "c72a6e9b-0423-4260-b6d0-e21192d78fa8"
                            }
                        },
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "d46d8ac6-55c8-4097-8152-c4507a7744bb"
                            }
                        },
                        {
                            "dataset": {
                                "enabled": True,
                                "id": "030df7b2-30c9-4be2-bec2-44fdde5a9244"
                            }
                        }
                    ]
                },
                "metadata_filtering_mode": "disabled",
                "metadata_model_config": {
                    "provider": "langgenius/tongyi/tongyi",
                    "name": "qwen2.5-72b-instruct",
                    "mode": "chat",
                    "completion_params": {
                        "temperature": 0.7
                    }
                }
            },
            "file_upload": {
                "image": {
                    "detail": "high",
                    "enabled": False,
                    "number_limits": 3,
                    "transfer_methods": [
                        "remote_url",
                        "local_file"
                    ]
                },
                "enabled": False,
                "allowed_file_types": [],
                "allowed_file_extensions": [
                    ".JPG",
                    ".JPEG",
                    ".PNG",
                    ".GIF",
                    ".WEBP",
                    ".SVG",
                    ".MP4",
                    ".MOV",
                    ".MPEG",
                    ".WEBM"
                ],
                "allowed_file_upload_methods": [
                    "remote_url",
                    "local_file"
                ],
                "number_limits": 3,
                "fileUploadConfig": {
                    "file_size_limit": 15,
                    "batch_count_limit": 5,
                    "image_file_size_limit": 10,
                    "video_file_size_limit": 100,
                    "audio_file_size_limit": 50,
                    "workflow_file_upload_limit": 10
                }
            },
            "annotation_reply": {
                "enabled": False
            },
            "supportAnnotation": True,
            "appId": "4305c12e-60c4-4fca-bca2-ae81043854aa",
            "supportCitationHitInfo": True,
            "model": {
                "provider": "langgenius/gemini/google",
                "name": "gemini-2.5-flash",
                "mode": "chat",
                "completion_params": {
                    "stop": [],
                    "include_thoughts": False,
                    "thinking_mode": False
                }
            }
        },
        "parent_message_id": None,
        "conversation_id": conversation_id
    }

    try:
        print(url)
        resp = requests.post(url, headers=headers, json=payload, timeout=60)
        resp.raise_for_status()
        # 关键点 2: 使用 iter_lines() 逐行读取
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    print(event_data)
                    try:
                        event_json = json.loads(event_data)
                        conversation_id = event_json.get("conversation_id")
                        if conversation_id is not None:
                            print(conversation_id)
                            return conversation_id
                    except json.JSONDecodeError:
                        continue



    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        # print(resp.json())
        # print(resp.text)
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None



# 评估答案函数
def evaluate_answer(question: str, answer: str, standard_answer:str) -> str:
    print("_____________________________________")
    print("开始评估答案")
    user_prompt = f"""
                请你扮演一个 FAQ 质量评估专家，根据以下输入判断 实际回答 的质量。
                注意：标准答案不一定是完全正确的，只作为参考信息，不是最终真理。

                🧩 输入格式

                    用户问题：{question}

                    标准答案（参考）：{standard_answer}

                    实际回答（待评估）：{answer}

                ✅ 评估任务

                请从以下角度进行分析：

                1. 相关性（0–100）

                    实际回答与用户问题是否相关？是否切题？

                2. 准确性（0–100）

                    在没有绝对正确答案时，请根据常识、逻辑、行业知识进行判断。
                    不依赖标准答案，而是独立思考实际回答是否合理、事实是否正确。

                3. 与标准答案的一致性（0–100）

                    比较实际回答 vs 标准答案

                    内容是否一致

                    是否含义相同

                    是否补充了没提到的信息

                    但注意，不要认为标准答案一定对，只用于对比。

                ✅ 输出要求
                    请按以下结构输出：

                    【评分】
                    相关性：X/100
                    准确性：X/100
                    与标准答案一致性：X/100

                    【差异分析】
                    - 实际回答与标准答案的主要差异点是哪些？
                    - 是否存在冲突？如果有，指出冲突原因。
                    - 哪个答案更合理（并解释原因）？

                    【改进建议】
                    给出 1–3 条实际回答的优化建议。

                ✅ 示例（可选）
                    你不需要使用下面示例，只需理解格式即可：
                    【评分】
                    相关性：85
                    准确性：70
                    一致性：60

                    【差异分析】
                    - 实际回答更多强调流程，而标准答案强调规则。
                    - 存在部分冲突，例如标准答案称“不需要准备文件”，但实际回答要求“身份证明”。

                    【改进建议】
                    1. 用更明确的步骤表达流程
                    2. 避免和标准答案冲突的说法
                    3. 明确身份验证是否必需
                """


    system_message = "你是语意判断专家"
    return chat_dify_llm(system_message, user_prompt, token)


# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False
# 主处理流程
def process_excel(input_file, output_file):
    # 读取Excel文件
    try:
        df = pd.read_excel(input_file)
        print(f"成功读取Excel文件，共{len(df)}条记录")
        print(f"Excel文件列名：{list(df.columns)}")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return

    # 确保存在问题列
    if '问题' not in df.columns:
        print("Excel文件中缺少'问题'列")
        return

    # 创建新列（如果不存在）
    if 'question_classifier' not in df.columns:
        df['question_classifier'] = ""
    if 'knowledge_retrieval' not in df.columns:
        df['knowledge_retrieval'] = ""

    # 处理每个问题
    index_start = 0

    round = 15
    conversation_id = None
    for index, row in df.iloc[index_start:].iterrows():
        # if index % 16 == 0:
        #     token = login()
        # 每16轮重新获取token
        # if round % 16 == 15:
        #     token = login()

        round += 1
        question = str(row['问题']).strip()
        standard_answer = str(row['标准答案']).strip()
        if not question or question.lower() == 'nan' or question == '' or question == '-':
            continue

        print(f"\n处理问题 {index + 1}/{len(df)}: {question}")

        # TODO：调用chatflow发送请求
        conversation_id = chat_chatlow(question, conversation_id, token)
        print(f"conversation_id: {conversation_id}")
        # TODO：调用chatflow_conversation获取answer
        answer, knowledge_retrieval = chatflow_conversation(conversation_id, token)
        if not answer:
            print(f"获取答案失败，跳过问题: {question}")
            continue
        # TODO: 调用llm判断问题答案
        # 评估答案
        evaluation = evaluate_answer(question, answer, standard_answer)
        if not evaluation:
            print(f"评估答案失败，跳过问题: {question}")
            continue
        print(f"评估结果: {evaluation}")
        # 更新DataFrame
        df.at[index, '答案'] = answer
        # df.at[index, 'question_classifier'] = question_classifier
        df.at[index, 'knowledge_retrieval'] = knowledge_retrieval
        df.at[index, '评估结果'] = evaluation
        # print(f"工作流的回答: {answer}")
        # print(f"评测结果: {evaluation}")
        print("="*200)

        # 每处理3条保存一次进度
        if (index + 1) % 3 == 0:
            # 保存到临时文件
            temp_file = f"temp_{output_file}"
            df.to_excel(temp_file, index=False)

            # 格式化临时文件
            if format_excel(temp_file):
                # 重命名为最终文件
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"已保存格式化进度到 {output_file}")
            else:
                print("格式化失败，保留未格式化文件")

    # 最终保存结果
    try:
        # 保存到临时文件
        temp_file = f"temp_{output_file}"
        df.to_excel(temp_file, index=False)

        # 格式化并重命名
        if format_excel(temp_file):
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！格式化结果已保存到 {output_file}")
        else:
            # 如果格式化失败，保留未格式化文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！未格式化结果已保存到 {output_file}")

        print(f"共处理 {len(df)} 条问题，成功 {df['答案'].notnull().sum()} 条")
    except Exception as e:
        print(f"保存结果失败: {e}")

# 使用示例
if __name__ == "__main__":
    # input_excel = "question-workshop.xlsx"  # 输入文件名
    # output_excel = "output-workshop-0805.xlsx"  # 输出文件名
    # input_excel = "question-orginalRM.xlsx"  # 输入文件名
    # output_excel = "output_orginalRM_0806.xlsx"  # 输出文件名
    # input_excel = "D:/codeWorkspace/trae/automation-test/tests/QA_2/Q_related-FUSE_FINA.xlsx"  # 输入文件名（使用正斜杠避免转义问题）
    input_excel = "Q_related-FUSE_FINA_1.xlsx"  # 输入文件名（使用正斜杠避免转义问题）

    output_excel = "A_related-FUSE_FINA_monica_1112_1.xlsx"  # 输出文件名
    # input_excel = "Q_IntentDetect_Sample.xlsx"  # 输入文件名
    # output_excel = "A_IntentDetect_Sample.xlsx"  # 输出文件名
    # input_excel = "Q_Claim_condition.xlsx"  # 输入文件名
    # output_excel = "A_Claim_condition.xlsx"  # 输出文件名
    process_excel(input_excel, output_excel)

