import os
from time import sleep
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *


base_url = sit_base_url
email = sit_email
password =sit_password
workflow_app_id =sit_workflow_app_id
llm_app_id = sit_llm_app_id
webhook_url = sit_webhook_url
wa_id = whatsapp_id


# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]

# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False

# 主处理流程
def process_excel(input_file, output_file):
    # 读取Excel文件
    try:
        df = pd.read_excel(input_file)
        print(f"成功读取Excel文件，共{len(df)}条记录")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return

    # 确保存在问题列
    if 'mobile' not in df.columns:
        print("Excel文件中缺少'mobile'列")
        return


    # 处理每个问题
    index_start = 1

    for index, row in df.iloc[index_start:].iterrows():

        mobile = str(row['mobile']).strip()
        email  = str(row['email']).strip()
        nickName = str(row['nickName']).strip()


        if mobile == '' or email == '' or nickName == '':
            continue

        print(f"\n处理{nickName} /{email}: {mobile}")

        # 调用注册接口
        result = request_Register(nickName, email, mobile)

        print("=" * 200)

        # # 保存答案
        # if answer is None:
        #     raise ValueError("数据库没有找到答案")
        #
        # output_answer += answer

        # 更新DataFrame
        df.at[index, 'result'] = result



        # 每处理3条保存一次进度
        if (index + 1) % 3 == 0:
            # 保存到临时文件
            temp_file = f"temp_{output_file}"
            df.to_excel(temp_file, index=False)

            # 格式化临时文件
            if format_excel(temp_file):
                # 重命名为最终文件
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"已保存格式化进度到 {output_file}")
            else:
                print("格式化失败，保留未格式化文件")

    # 最终保存结果
    try:
        # 保存到临时文件
        temp_file = f"temp_{output_file}"
        df.to_excel(temp_file, index=False)

        # 格式化并重命名
        if format_excel(temp_file):
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！格式化结果已保存到 {output_file}")
        else:
            # 如果格式化失败，保留未格式化文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！未格式化结果已保存到 {output_file}")

        print(f"共处理 {len(df)} 条问题，成功 {df['答案'].notnull().sum()} 条")
    except Exception as e:
        print(f"保存结果失败: {e}")



def requests_otp(mobile):
    print("1.1 Sending OTP request...")
    url = f"https://ptr-uat.fuse.co.id/api/prm/verificationCode/send"
    headers = {
        "content-type": "application/json",
        "clientType": "ANDROID",
        "appCode": "IDP_FUSE_PRO"
    }
    # if type(msg_body) != str:
    #     query = str(msg_body)
    # user_input="reset session"

    if mobile and mobile == '':
        return "Mobile为空"


    payload = {
        "address": mobile,
        "type": "REGISTER_PARTNER",
        "tenantId": "1000662"
    }

    try:
        resp = requests.post(url, headers=headers, json=payload)
        # print(resp)
        # resp.raise_for_status()
        print(resp.text)
        print(resp.json())

    except requests.RequestException as e:
        print(f"Error calling requests_otp: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None
    print(resp.json()["header"]["message"])
    print(f"request otp success {resp.json()['header']['resultCode']}")

    return resp.json()["header"]["resultCode"]


def queryDB_byMobile(mobile):
    print("1.2 start queryDB_byMobile")
    # sleep(10)
    url = f"https://rd-dms.fuseinsurtech.com/query/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/x-www-form-urlencoded",
        "cookie": "csrftoken=qRP2ywXLerNNjkggzaIJTGwJBsMvIDNSVywD0wv1BOrTo9E806cHs4SkYHg3l3z0; sessionid=r632jwxeb4qk7wadyliijdpsv71kcm4y",
        "x-csrftoken": "qRP2ywXLerNNjkggzaIJTGwJBsMvIDNSVywD0wv1BOrTo9E806cHs4SkYHg3l3z0"
    }
    if type(mobile) != str:
        mobile = str(mobile)

    # trace_id = '8c6f64a18338a8da'

    sql_content = f"select data->'$.verificationCade' from message.sms_record where mobile='{mobile}' order by id desc limit 1"

    print(sql_content)
    # TODO：动态获取messgae from的wa_id, wa_token，business_account_id
    payload = {
        "instance_name": "ID_UAT_CORE_MYSQL8.0",
        "db_name": "message",
        "schema_name": None,
        "tb_name": None,
        "sql_content": sql_content,
        "limit_num": "100"
    }
    try:
        resp = requests.post(url, headers=headers, data=payload)
        resp = requests.post(url, headers=headers, data=payload)
        print(resp.json())

    except requests.RequestException as e:
        print(f"Error calling db: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

    if resp.json()['data']['affected_rows'] == 0:
        print("no verification code")
        result = None
    else:
        print(resp.json()['data']['rows'][0][0])
        result = resp.json()['data']['rows'][0][0].strip('"')
        print(result)
        print("end request_Register")
    return result


def request_Register(nickName, email, mobile):
    print("start request_Register")
    """
    调用webhook发送文本消息
    """
    # https://pchat-uat.fuse.co.id/api/ai/chatbot/whatsapp/webhook

    url = f"https://ptr-uat.fuse.co.id/api/prm/partner/register"
    headers = {
        "content-type": "application/json",
        "clientType" : "ANDROID",
        "appCode" : "IDP_FUSE_PRO"
    }
    # if type(msg_body) != str:
    #     query = str(msg_body)
    # user_input="reset session"

    if mobile and mobile != '':
        # 触发OTP
        resultCode = requests_otp(mobile)
        if resultCode == 200:
            # 调用数据库接口，查询OTP
            otp = queryDB_byMobile(mobile)
        else:
            return None

    payload = {
                "idd": "62",
                # "mobile": "8123121212",
                "mobile": mobile,
                "companyName": "",
                "nickname": nickName,
                "password": "123456a",
                "occupationUid": 1282,
                "occupationCustom": "",
                "code1": "1",
                "code2": "2",
                "code3": "3",
                "code4": "4",
                "accountType": 1,
                "CORPORATE_ARTICLE_OF_ASSOCIATION": [],
                "CORPORATE_PERMISSION_DOCUMENT": [],
                "CORPORATE_PIC_KTP": [],
                "CORPORATE_TAX_DOCUMENT": [],
                "CORPORATE_POWER_OF_ATTORNEY": [],
                "email": email,
                "inviteFuseId": "99485f35",
                "c": "JnRlbmFudElkPTEwMDA2NjImbGlua0Zvcm09MSZjaGFubmVsPVNoYXJl",
                "region": "id",
                "utm_source": "wa",
                "lan": "id",
                "channel": "Share",
                "tenantId": "1000662",
                "linkForm": "1",
                "verificationCode": otp,
                "inviteType": 1,
                "language": "in_ID",
                "attachmentList": []
            }

    print(payload)
    try:
        resp = requests.post(url, headers=headers, json=payload)
        # print(resp)
        # resp.raise_for_status()
        print(resp.text)

    except requests.RequestException as e:
        print(f"Error calling webhook: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

    return resp.text


def getAnswerFromDB_bytraceId(trace_id):
    print("start queryDB_bytraceId")
    # 休眠5s
    sleep(10)
    """
    调用webhook发送文本消息
    """
    # https://pchat-uat.fuse.co.id/api/ai/chatbot/whatsapp/webhook

    url = f"https://rd-dms.fuseinsurtech.com/query/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/x-www-form-urlencoded",
        "cookie": "csrftoken=PVbSL3lxBK6fj6APFze7qkOOJ8PsQKzH3GC4O8GyWX1w6IGz88c4y57Fth8wfV94; sessionid=mmpqqm0ofcm3pl7p3ta795bksljgo0pv",
        "x-csrftoken": "PVbSL3lxBK6fj6APFze7qkOOJ8PsQKzH3GC4O8GyWX1w6IGz88c4y57Fth8wfV94"
    }
    if type(trace_id) != str:
        trace_id = str(trace_id)

    # trace_id = '8c6f64a18338a8da'

    sql_content = f"select reply from message.whatsapp_chat_record where 1=1 and trace_id = '{trace_id}' order by uid desc limit 1"

    # TODO：动态获取messgae from的wa_id, wa_token，business_account_id
    payload = {
        "instance_name" : "SG_SIT_MYSQL8_RW",
        "db_name" : "message",
        "schema_name" : None,
        "tb_name": None,
        "sql_content" : sql_content,
        "limit_num" : "100"
    }
    try:
        resp = requests.post(url, headers=headers, data=payload)
        resp = requests.post(url, headers=headers, data=payload)
        print(resp.json())

    except requests.RequestException as e:
        print(f"Error calling db: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

    if resp.json()['data']['affected_rows'] == 0:
        answer = trace_id + "在数据库中没有找到数据"
    else:
        answer = resp.json()['data']['rows'][0][0]

    print(answer)
    return answer

# 使用示例
if __name__ == "__main__":
    input_excel = "batchRegister.xlsx"  # 输入文件名
    output_excel = "result_batchRegister.xlsx"  # 输出文件名
    process_excel(input_excel, output_excel)

