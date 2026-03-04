import os
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *
from time import sleep

env = 'sit'
base_url = sit_base_url
email = sit_email
password =sit_password
llm_app_id = sit_sf_chatflow_intentDetect_app_id

# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]

# 调用模型函数
def chat_dify_llm_node(query: str, token):
    # token = login()
    app_id = llm_app_id
    # https://rd-dify-sit.fuse.co.id/console/api/apps/5749f5da-9fc2-463d-b00e-9893cc290b9b/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{app_id}/workflows/draft/nodes/llm/run"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {token}",
        "content-type": "application/json",
    }

    payload = {"inputs": {"#sys.query#": query}}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        print(resp.json())

        if resp.json()['status'] == 'failed':
            print(f"Error calling chat: {resp.json()['message']}")
            return None
        else:
            print(resp.json()['outputs']['text'])
            return resp.json()['outputs']['text']

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")

    return None

# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False
# 主处理流程
def process_excel(input_file, output_file):
    # 读取Excel文件
    try:
        df = pd.read_excel(input_file)
        print(f"成功读取Excel文件，共{len(df)}条记录")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return

    # 确保存在问题列
    if 'eng' not in df.columns:
        print("Excel文件中缺少'eng问题'列")
        return

    # 创建新列（如果不存在）
    if '答案' not in df.columns:
        df['答案'] = ""
    if '评估结果' not in df.columns:
        df['评估结果'] = ""

    # 处理每个问题
    index_start = 0

    round = 15
    for index, row in df.iloc[index_start:].iterrows():
        # if index % 16 == 0:
        #     token = login()
        # 每16轮重新获取token
        if round % 16 == 15:
            token = login()

        round += 1
        question = str(row['eng']).strip()
        if not question or question.lower() == 'nan' or question == '' or question == '-':
            continue

        print(f"\n处理问题 {question}")

        # 获取答案和知识库召回结果
        result = chat_dify_llm_node( question, token)
        if result is not None:
            answer = result
        else:
            # 处理 None 的情况，例如抛出异常或返回默认值
            raise ValueError("chat_workflow 返回了 None")

        if not answer:
            print(f"获取答案失败，跳过问题: {question}")
            continue


        # 更新DataFrame
        df.at[index, '答案_eng'] = result


        question = str(row['Bahasa']).strip()
        if not question or question.lower() == 'nan' or question == '' or question == '-':
            continue

        print(f"\n处理问题 {question}")

        # 获取答案和知识库召回结果
        result = chat_dify_llm_node( question, token)
        if result is not None:
            answer = result
        else:
            # 处理 None 的情况，例如抛出异常或返回默认值
            raise ValueError("chat_workflow 返回了 None")

        if not answer:
            print(f"获取答案失败，跳过问题: {question}")
            continue


        # 更新DataFrame
        df.at[index, '答案_Bahasa'] = result

        print("="*200)

        sleep(3)

        # 每处理3条保存一次进度
        if (index + 1) % 3 == 0:
            # 保存到临时文件
            temp_file = f"temp_{output_file}"
            df.to_excel(temp_file, index=False)

            # 格式化临时文件
            if format_excel(temp_file):
                # 重命名为最终文件
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"已保存格式化进度到 {output_file}")
            else:
                print("格式化失败，保留未格式化文件")

    # 最终保存结果
    try:
        # 保存到临时文件
        temp_file = f"temp_{output_file}"
        df.to_excel(temp_file, index=False)

        # 格式化并重命名
        if format_excel(temp_file):
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！格式化结果已保存到 {output_file}")
        else:
            # 如果格式化失败，保留未格式化文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！未格式化结果已保存到 {output_file}")

        print(f"共处理 {len(df)} 条问题，成功 {df['答案'].notnull().sum()} 条")
    except Exception as e:
        print(f"保存结果失败: {e}")

# 使用示例
if __name__ == "__main__":
    # input_excel = "question-workshop.xlsx"  # 输入文件名
    # output_excel = "output-workshop-0805.xlsx"  # 输出文件名
    # input_excel = "question-orginalRM.xlsx"  # 输入文件名
    # output_excel = "output_orginalRM_0806.xlsx"  # 输出文件名
    input_excel = "question.xlsx"  # 输入文件名
    output_excel = "output-0818_uat.xlsx"  # 输出文件名
    # input_excel = "Q_意图识别样本.xlsx"  # 输入文件名
    # output_excel = "A_意图识别样本.xlsx"  # 输出文件名
    process_excel(input_excel, output_excel)

