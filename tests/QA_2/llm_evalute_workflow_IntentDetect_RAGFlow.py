'''
    调用dify rag workflow获取意图识别结果、知识库召回结果、最后输出答案
    1. 建议：
     运行前reset, 然后获取到新的session_id，填写到下面的字段

    
'''

import os
import time
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *
from fileProcess import FileProcess
import re
from typing import Dict, Any, Optional
from multiprocessing import Pool

env = 'sit'
base_url = uat_base_url
email = uat_email
password =uat_password
workflow_app_id = uat_workflow_app_id
llm_app_id = uat_llm_app_id
chat_flow_id = uat_sf_chatflow_id
session_id = 'SN00OL1WWIK000000003U'
domain = prod_domain

# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]


# 调用模型函数
def chat_dify_llm(system: str, query: str, token):
    # token = login()
    app_id = llm_app_id
    # https://rd-dify-sit.fuse.co.id/console/api/apps/5749f5da-9fc2-463d-b00e-9893cc290b9b/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{app_id}/workflows/draft/run"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {token}",
        # "authorization": f"Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }

    payload = {"inputs":{"user":query,"system":system},"files":[]}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == "workflow_finished":
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("text", final_answer)
                    except json.JSONDecodeError:
                        continue

        if final_answer is not None:
            return final_answer
        else:
            print("No valid answer found in SSE stream.")
            return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

def chat_workflow(query: str, access_token:str) -> tuple[Any | None, Any | None, Any | None] | None:
    print("enter chat_workflow")
    """
    对指定数据集做 hit-testing，返回接口完整 JSON。
    retrieval_model 留空时使用默认 hybrid_search 配置。
    """
    # https://rd-dify-sit.fuse.co.id/console/api/apps/a1936725-1498-452c-8310-b81e94936703/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{workflow_app_id}/workflows/draft/run"
    print(url)
    # access_token = login()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {access_token}",
        # "authorization": f"Bearer logineyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
        "sec-fetch-site": "same-origin",
        "sec-fetch-mode": "cors"
    }
    if type(query) != str:
        query=str(query)
    # query="Do you have car product?"


    print("workflow_question: ", query)
    payload = {
        "inputs": {
            "question": query,
            # "session_id": session_id,
            "domain": domain,
            "lan": "ID"
        },
        "files": []
    }

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        knowledge_retrieval = None
        question_classifier = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    print(event_data)
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == 'node_finished':
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                            # 获取检索到的知识库，process_guidance和general Qa知识库节点是知识库结果聚合：1753256233555；修车店是1753844464410
                            if event_json.get("data", {}).get("node_id") == '1753256233555':
                                knowledge_retrieval = event_json.get("data", {}).get("outputs", {}).get("output")
                                print("knowledge_retrieval in workflow: ", knowledge_retrieval)
                            elif event_json.get("data", {}).get("node_id") == '1753666462887':
                                knowledge_retrieval = event_json.get("data", {}).get("outputs", {}).get("result")
                                print("knowledge_retrieval in workflow: ", knowledge_retrieval)
                            # 获取意图分类结果，有两个节点
                            # 第一次意图识别，不带上下文，node id = 17553530659570
                            elif event_json.get("data", {}).get("node_id") == '17553530659570':
                                question_classifier = event_json.get("data", {}).get("outputs", {}).get("text")
                                print("RAG第一次意图识别结果: ", question_classifier)
                            # 第二次意图识别，带上下文，node id = 17553530659570
                            elif event_json.get("data", {}).get("node_id") == '1752731767860':
                                question_classifier = event_json.get("data", {}).get("outputs", {}).get("text")
                                print("RAG第二次意图识别结果: ", question_classifier)
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("answer", final_answer)
                    except json.JSONDecodeError:
                        continue

        print("final_answer", final_answer)
        print("question_classifier", question_classifier)
        print("knowledge_retrieval", knowledge_retrieval)

        # 使用 json.dumps() 将字典转换为 JSON 格式的字符
        knowledge_retrieval = json.dumps(knowledge_retrieval)

        # if final_answer is not None and question_classifier is not None and knowledge_retrieval is not None:
        #     print("knowledge_retrieval："+knowledge_retrieval)
        #     return final_answer, question_classifier, knowledge_retrieval
        # else:
        #     print("No valid answer found in SSE stream.")
        #     return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

    return final_answer, question_classifier, knowledge_retrieval

def chat_chatlow(query: str, access_token:str) -> tuple[Any | None, Any | None, Any | None, Any | None, Any | None] | None:
    print("enter chat_workflow")
    """
    对指定数据集做 hit-testing，返回接口完整 JSON。
    retrieval_model 留空时使用默认 hybrid_search 配置。
    """

    # https://rd-dify-sit.fuse.co.id/console/api/apps/f71f507b-da44-4d92-828b-6a37bf4d96a7/advanced-chat/workflows/draft/run
    # https://rd-dify-sit.fuse.co.id/console/api/apps/a1936725-1498-452c-8310-b81e94936703/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{chat_flow_id}/advanced-chat/workflows/draft/run"
    # access_token = login()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {access_token}",
        # "authorization": f"Bearer logineyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }
    if type(query) != str:
        query=str(query)
    # query="Do you have car product?"

    payload = {
            "files": [],
            "inputs": {
                "chat_temp_token": "eyJhbGciOiJIUzI1NiIsInppcCI6IkRFRiJ9.eNpEjMsOgjAQRf-la0ha6LTgWkOMxkeEhcu2jBEfpeERjcZ_d2TjrCbn3HvfrB8tmzF3NkOJ91C2V_QsYsF0g8duG9Ava_Jr2MFG71eLopgfNQWo4D3eqh67KZApkWcpl1JIJcjb5nVox84huQeN9yaE_27VUEdwzpVK-O_ozyKGz0BYg9ZagpYRa8wwARBZMoHL0NCgce7kRAKxMtbEUqQitoAuljJFa3JMawD2-QIAAP__.v9KsrJFHoaT-aIZPDQJHxB9iOlJey1icQvVKMGC6FVs",
                "partner_uid": "1000662000001008",
                "tenant_id": "1000662",
                "channel_user_id": "8619830441461",
                "channel_type": "whatsapp",
                "session_id": session_id,
                "lan": "ID"
            },
            "query": query,
            # "conversation_id": "d0c0b359-f6d6-4308-bab9-4559aa35f29a",
            # "parent_message_id": "c66fc8ff-1453-4b55-80ac-2b1802324cc4"
        }

    # 定义transation意图的枚举值列
    transation = ['INTENT_QUICK_ORDER', 'INTENT_QUOTE_COMPARISON', 'INTENT_ORDER', 'INTENT_PAYMENT', 'INTENT_INVITE', 'INTENT_WITHDRAWAL', 'INTENT_COMPARISON_LIST']




    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        knowledge_retrieval = None
        Level1_question_classifier = None
        question_classifier = None
        question_rephase = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    print(event_data)
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == 'node_finished':
                            if event_json.get("data", {}).get("node_id") == '1753082979250':    # "代码执行：意图识别结果ETL
                                Level1_question_classifier = event_json.get("data", {}).get("outputs", {}).get("inent_result")
                            elif event_json.get("data", {}).get("node_id") == '17552427857160':    # LLM：意图识别与拆分重写
                                question_rephase = event_json.get("data", {}).get("outputs", {}).get("text")
                    except json.JSONDecodeError:
                        continue

        print("一级意图分类: ", Level1_question_classifier)
        if Level1_question_classifier not in transation:
            #调用RAG workflow
            result = chat_workflow(query, access_token)
            if result is not None:
                final_answer, question_classifier, knowledge_retrieval = result
                print("一级意图分类: ", question_classifier)
            else:
                question_classifier = None

        print("最后结果————————————————————————")
        print("final_answer", final_answer)
        print("Level1_question_classifier", Level1_question_classifier)
        print("question_classifier", question_classifier)
        print("knowledge_retrieval", knowledge_retrieval)
        if final_answer == None:
            final_answer = " "
        if question_rephase == None:
            question_rephase = " "
        if Level1_question_classifier == None:
            Level1_question_classifier = " "
        if question_classifier == None:
            question_classifier = " "
        if knowledge_retrieval == None:
            knowledge_retrieval = " "



        return final_answer, question_rephase, Level1_question_classifier, question_classifier, knowledge_retrieval



        # 使用 json.dumps() 将字典转换为 JSON 格式的字符
        knowledge_retrieval = json.dumps(knowledge_retrieval)


    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None



# 评估答案函数
def evalsite_answer(question: str, llm_answer: str,token:str, right_answer:str) -> str:
    user_prompt = f"""
                # Role
                你是一名资深的保险专家，你的任务是对另一个AI模型（Gemini Flash）生成的答案进行严格、公正的质量评估。
                
                ## 评估背景
                - **待评估的问题：** “{question}”
                - **待评估的答案（来自被测试模型如Gemini Flash）：** “{llm_answer}”
                - **标准答案（如有，请提供）：** “{right_answer}”
                
                ## 核心指令
                请根据是否提供了“标准答案”来决定评估模式：
                - **【模式一：客观比对模式】**：如果提供了明确的标准答案，请严格比对待评估答案与标准答案的一致性。
                - **【模式二：主观评估模式】**：如果未提供标准答案（标注为‘无’），请基于通用高质量回答标准进行评估。
                
                ## 评估流程与标准
                               
                ### 【模式一：客观比对模式】（当有标准答案时）
                请从以下三个维度评估待评估答案与标准答案的吻合度（总分100分）：
                - **相关性 (30分)**：答案是否直接针对问题，是否答非所问？
                - **准确性 (50分)**：答案中的事实、数据、概念是否与标准答案完全一致？是否存在错误？
                - **完整性 (20分)**：是否涵盖了标准答案中的所有关键要点？是否有重大遗漏？
                
                ### 【模式二：主观评估模式】（当无标准答案时）
                从以下五个核心维度对上述答案进行逐一分析和评分（每个维度满分20分，总分100分）。
                - **a. 事实准确性与专业性 (20分)**
                  - 答案中的保险术语（如免赔额、现金价值、受益人、责任免除等）使用是否准确？
                  - 所述保险知识、原理、法律法规（如《保险法》相关原则）是否正确？
                  - 是否存在事实性错误或过时信息？
                - **b. 逻辑严谨性与完整性 (20分)**
                  - 推理过程是否清晰、有逻辑？能否清晰地解释“为什么”？
                  - 是否涵盖了问题中所有关键点？是否考虑了不同的情景或例外情况（如免责条款）？
                  - 答案结构是否完整（如：先解释概念，再分析案例，最后总结）？
                - **c. 清晰度与可读性 (20分)**
                  - 答案是否组织良好、易于理解？是否使用了分点、分段等格式？
                  - 能否用通俗的语言向非专业人士解释复杂概念？
                  - 语言是否流畅、简洁，没有不必要的赘述？
                - **d. 谨慎性与负责任程度 (20分)**
                  - 答案是否包含了必要的免责声明（如“具体请以保险合同条款为准”、“建议咨询专业的保险顾问或保险公司”）？
                  - 是否避免了给出绝对的、可能构成医疗或法律建议的结论？
                  - 态度是否中立、客观，不会误导用户？
                - **e. 相关性与实用性 (20分)**
                  - 答案是否直接回答了核心问题，没有答非所问或避重就轻？
                  - 对于用户提出的问题，给出的信息是否具有实际操作价值？能否真正帮助用户解决疑惑或做出下一步决策？
                                  
                ## 输出格式要求
                请严格按照以下格式组织你的最终评估报告：
                
                **【评估模式】**
                [在此声明采用的评估模式：模式一（客观比对）或 模式二（主观评估）]
                
                **【总体评分】**
                [总分] / 100
                
                **【分维度评分】**
                - [维度1名称]: [得分] / [满分]
                - [维度2名称]: [得分] / [满分]
                - [维度3名称]: [得分] / [满分]
                - [维度4名称（如有）]: [得分] / [满分]
                
                **【详细分析】**
                - **优点：** 列出待评估答案中表现突出的部分。
                - **缺点与错误：** 列出待评估答案中存在的具体问题、错误或遗漏。在模式一下，必须明确指出与标准答案不符之处。
                - **改进建议：** 提供1-2句具体的优化建议。
                
                # 特别注意：
                # 1. 用中文输出评估结果
                """


    system_message = "你是语意判断专家"
    return chat_dify_llm(system_message, user_prompt, token)


# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False



# 主处理流程
def process_excel(input_file, output_file):

    # 读取文件
    df = FileProcess.real_excel(input_file)

    # 处理每个问题
    index_start = 0

    round = 15
    for index, row in df.iloc[index_start:].iterrows():
        # 每16轮重新获取token
        if round % 16 == 15:
            token = login()

        round += 1
        question = str(row['问题']).strip()
        # 暂时没有多轮对话给标答的情况不拆分
        standard_answer = str(row['标准答案'])
        if not question or question.lower() == 'nan' or question == '' or question == '-':
            continue

        print(f"\n处理问题 {index + 1}/{len(df)}: {question}")

        knowledge_highestScore = None

        # 2. 调用chatflow处理问题
        # 准备多进程参数
        args_list = [(token, index, question, standard_answer, df, output_file)]

        # 使用多进程池处理文件
        with Pool(processes=1) as pool:
            pool.map(process_one_question, args_list)
        #
        #
        #
        # output_answer = ''
        # output_question_rephase= ''
        # output_level1_question_classifier = ''
        # output_knowledge_retrieval = ''
        # output_question_classifier = ''
        # output_knowledge_highestScore = ''
        # output_evalsition = ''
        #
        # # 2.1 分解这个问题的每个步骤
        # steps = [q for q in re.split(r'\d+\.', question) if q]
        # print("全部步骤：", steps)
        #
        #
        # # 2.2 逐个步骤调用chatflow / workflow
        # for step_index in range(0, len(steps)):
        #     # 初始化全部变量
        #     answer, question_rephase, Level1_question_classifier, question_classifier, knowledge_retrieval, knowledge_highestScore = '', '', '', '', '', ''
        #
        #     print(f"处理步骤{step_index+1}/{steps.__len__()}: {steps[step_index]}")
        #     # output_chatlog += f"user: {step_index}"
        #
        #     # 获取答案和知识库召回结果
        #     result = chat_workflow(steps[step_index], token)
        #
        #     # if result is not No
        #     if result is not None:
        #         answer, question_classifier, knowledge_retrieval = result
        #     else:
        #         # 处理 None 的情况，例如抛出异常或返回默认值
        #         print("chat_workflow 返回了 None")
        #         continue
        #
        #     print("check knowledge_retrieval before json.loads: ", knowledge_retrieval)
        #     # 如果knowledge_retrieval 不等于Null, 将knowledge_retrieval转换成列表
        #     if knowledge_retrieval is None:
        #         knowledge_highestScore = "None"
        #     elif knowledge_retrieval is not None:
        #         knowledge_retrieval_temp = json.loads(knowledge_retrieval)
        #         print("knowledge_retrieval_temp: ", knowledge_retrieval_temp)
        #         if knowledge_retrieval_temp and len(knowledge_retrieval_temp) > 0:
        #             #  获取列表第一条，最高分的知识
        #             knowledge_highestScore = knowledge_retrieval_temp[0]
        #             # 转换成字符串
        #             knowledge_highestScore = json.dumps(knowledge_highestScore)
        #             print("知识库最高分: ", knowledge_highestScore)
        #         else:
        #             knowledge_highestScore = "None"
        #
        #     # 评估答案
        #     evalsition = evalsite_answer(steps[step_index], answer, token, standard_answer)
        #     if not evalsition:
        #         print(f"评估答案失败，跳过问题: {question}")
        #         continue
        #     print(f"评测结果: {evalsition}")
        #     print("=" * 200)
        #
        #     # 保存答案
        #     # 答案
        #     output_answer += '-' + answer + '\n'
        #     # 问题拆分重写结果和意图识别优先级结果
        #     output_question_rephase += '-' + question_rephase + '\n'
        #     # chatflow意图分类结果
        #     output_level1_question_classifier += '-' + Level1_question_classifier + '\n'
        #     # 最终问题分类结果
        #     output_question_classifier += '-' + question_classifier + '\n'
        #     # 知识库检索结果
        #     output_knowledge_retrieval += '-' + knowledge_retrieval + '\n'
        #     # 知识库最高分
        #     output_knowledge_highestScore += '-' + knowledge_highestScore + '\n'
        #     # 评测结果
        #     output_evalsition += '-' + evalsition + '\n'
        #
        #
        # # 更新DataFrame
        # df.at[index, '答案'] = output_answer
        # df.at[index, '意图识别与拆分重写结果'] = output_question_rephase
        # df.at[index, 'Level1_question_classifier'] = output_level1_question_classifier
        # df.at[index, 'question_classifier'] = output_question_classifier
        # df.at[index, 'knowledge_retrieval'] = output_knowledge_retrieval
        # df.at[index, '最高分知识'] = output_knowledge_highestScore
        # df.at[index, '评估结果'] = output_evalsition
        # # print(f"评测结果: {evalsition}")
        # print("="*200)

        # 每处理3条保存一次进度
        FileProcess.save_result_temp(index= index, df= df, output_file=output_file)

    # 最终保存结果
    FileProcess.write_to_excel(df, output_file)





# 使用示例
if __name__ == "__main__":
    # input_excel = "question-workshop.xlsx"  # 输入文件名
    # output_excel = "output-workshop-0805.xlsx"  # 输出文件名
    # input_excel = "question-orginalRM.xlsx"  # 输入文件名
    # output_excel = "output_orginalRM_0806.xlsx"  # 输出文件名
    # input_excel = "Jess_SIT_IntentDetect.xlsx"  # 输入文件名
    # output_excel = "A_Jess_SIT_IntentDetect.xlsx"  # 输出文件名
    input_excel = "Q_related-FUSE_FINA.xlsx"  # 输入文件名
    output_excel = "A_related-FUSE_FINA_0916_sit_5.xlsx"  # 输出文件名
    # input_excel = "Q_Unsupported_producat_category.xlsx"  # 输入文件名
    # output_excel = "A_Unsupported_producat_category.xlsx"  # 输出文件名
    # input_excel = "Q_0901Fina Day Feedback.xlsx"  # 输入文件名
    # output_excel = "A_0901Fina Day Feedback_0903_2.xlsx"  # 输出文件名
    # input_excel = "Q_FinamustA_Failcase.xlsx"  # 输入文件名
    # output_excel = "A_FinamustA_Failcase_0829.xlsx"  # 输出文件名
    # input_excel = "fina day 意图识别case.xlsx"  # 输入文件名
    # output_excel = "A_fina day 意图识别case.xlsx"  # 输出文件名
    # input_excel = "sample意图测试.xlsx"  # 输入文件名
    # output_excel = "A_sample意图测试_0915_uat.xlsx"  # 输出文件名
    # input_excel = "Q_Claim_condition.xlsx"  # 输入文件名
    # output_excel = "A_Claim_condition.xlsx"  # 输出文件名
    # 打印开始执行的时间
    start_time = time.time()
    process_excel(input_excel, output_excel)
    # 打印结束时间
    end_time = time.time()
    print("开始执行时间：", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(start_time)))
    print("结束执行时间：", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(end_time)))
    print(f"总耗时: {end_time - start_time:.2f} 秒")


