'''
    调用dify chatflow获取一级意图识别结果和inforamtion侧二级意图识别结果
    1. 建议：
     运行前reset, 然后获取到新的session_id，填写到下面的字段

    
'''
import datetime
import os
import re
import time
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *
from fileProcess import  FileProcess
from tests.QA_2.llm_evalute_WebHook import evalsite_answer
from tests.QA_2.util.httpUtil import chat_chatlow
from tests.QA_2.util.llmEvaHandler import evalsite_answer_withStandAns

env = 'sit'
base_url = sit_base_url
email = sit_email
password =sit_password
workflow_app_id = sit_workflow_app_id
llm_app_id = sit_llm_app_id
chat_flow_id = sit_sf_v2
session_id = 'SN00OQUB02Z0000000041'
domain = prod_domain
user_info = sit_user_info

# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]



# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False
# 主处理流程



def process_excel(input_file, output_file):

    # 读取文件
    df = FileProcess.real_excel(input_file)

    # 处理每个问题
    index_start = 0

    round = 15
    for index, row in df.iloc[index_start:].iterrows():
        # 每16轮重新获取token
        if round % 16 == 15:
            token = login()

        round += 1

        # 获取当前行问题
        question = str(row['问题']).strip()
        # 获取当前行标准答案，暂时没有多轮对话给标答的情况不拆分
        standard_answer = str(row['标准答案'])

        if not question or question.lower() == 'nan' or question == '' or question == '-':
            continue

        print(f"\n处理问题 {index + 1}/{len(df)}: {question}")
        knowledge_highestScore = None

        # 2. 调用chatflow处理问题
        # 2.1 分解这个问题的每个步骤
        steps = [q for q in re.split(r'\d+\.', question) if q]
        print("全部步骤：", steps)

        output_answer = ''
        output_question_rephase= ''
        output_level1_question_classifier = ''
        output_knowledge_retrieval = ''
        output_question_classifier = ''
        output_evalsition = ''
        output_knowledge_highestScore = ''
        output_chatflow_final_output = ''
        output_duration_seconds = ''

        # 2.2 逐个步骤调用chatFlow / workflow
        for step_index in range(0, len(steps)):
            answer, chatFlow_final_output, duration_seconds = '', '', ''

            step_question = str(steps[step_index]).strip()
            print(f"处理步骤{step_index+1}/{steps.__len__()}: {step_question}")

            # 调用chatFlow
            result = chat_chatlow(base_url=base_url, app_id=chat_flow_id, userInfo=user_info, query=step_question, access_token=token)


            # if result is not No
            if result is not None:
                answer, chatFlow_final_output, duration_seconds = result
            else:
                # 处理 None 的情况，例如抛出异常或返回默认值
                print("chat_chatlow 返回了 None")
                answer, chatFlow_final_output, duration_seconds = ''
                continue

            # 如果answer = 空
            if answer == None:
                answer = ''
            if chatFlow_final_output == None:
                chatFlow_final_output = ''
            if duration_seconds == None:
                duration_seconds = 0


            # 评估答案 def evalsite_answer_withStandAns(base_url: str, app_id: str, question: str, llm_answer: str, token: str, right_answer: str)
            evalsition = evalsite_answer_withStandAns(base_url=base_url, app_id=llm_app_id, token=token, question=step_question, llm_answer=answer, right_answer=standard_answer)
            if not evalsition:
                print(f"评估答案失败，跳过问题: {question}")
                continue
            print(f"评测结果: {evalsition}")
            print("=" * 200)

            # 保存答案
            # 答案
            output_answer += f"{step_index+1}. " + answer + '\n'
            # chatflow最终输出
            output_chatflow_final_output += f"{step_index+1}. " + json.dumps(chatFlow_final_output) + '\n'
            # chatflow耗时
            output_duration_seconds += f"{step_index+1}. " + str(duration_seconds) + '\n'
            # output_evalsition += f"{step_index+1}. " + evalsition + '\n'
            # 评测结果
            output_evalsition += f"{step_index+1}. " + evalsition + '\n'


        # 更新DataFrame
        df.at[index, '答案'] = output_answer
        df.at[index, 'chatflow最终输出'] = output_chatflow_final_output
        df.at[index, 'chatflow耗时'] = output_duration_seconds
        df.at[index, '评估结果'] = output_evalsition
        print("="*200)

        # 每处理3条保存一次进度
        FileProcess.save_result_temp(index=index, df=df, output_file=output_file)

        # 最终保存结果
    FileProcess.write_to_excel(df, output_file)

# 使用示例
if __name__ == "__main__":
    # 打印开始执行的时间
    start_time = time.time()
    # input_excel = "question-workshop.xlsx"  # 输入文件名
    # output_excel = "output-workshop-0805.xlsx"  # 输出文件名
    # input_excel = "question-orginalRM.xlsx"  # 输入文件名
    # output_excel = "output_orginalRM_0806.xlsx"  # 输出文件名
    # input_excel = "Jess_SIT_IntentDetect.xlsx"  # 输入文件名
    # output_excel = "A_Jess_SIT_IntentDetect.xlsx"  # 输出文件名
    # input_excel = "All_Test_Case.xlsx"  # 输入文件名
    # output_excel = "Q_All_Test_Case_0918.xlsx"  # 输出文件名
    # input_excel = "Q_Unsupported_producat_category.xlsx"  # 输入文件名
    # output_excel = "A_Unsupported_producat_category.xlsx"  # 输出文件名
    # input_excel = "Q_0901Fina Day Feedback.xlsx"  # 输入文件名
    # output_excel = "A_0901Fina Day Feedback_0903_2.xlsx"  # 输出文件名
    # input_excel = "Q_FinamustA_Failcase.xlsx"  # 输入文件名
    # output_excel = "A_FinamustA_Failcase_0829.xlsx"  # 输出文件名
    # input_excel = "fina day 意图识别case.xlsx"  # 输入文件名
    # output_excel = "A_fina day 意图识别case.xlsx"  # 输出文件名
    # input_excel = "sample意图测试.xlsx"  # 输入文件名
    # output_excel = "A_sample意图测试_0915_uat.xlsx"  # 输出文件名
    # input_excel = "Q_Claim_condition.xlsx"  # 输入文件名
    # output_excel = "A_Claim_condition.xlsx"  # 输出文件名
    # input_excel = "Q_0901Fina Day Feedback.xlsx"  # 输入文件名
    input_excel = "Q_related-FUSE_FINA.xlsx"  # 输入文件名
    output_excel = "A_related-FUSE_FINA_multiRound_0930_1.xlsx"  # 输出文件名

    # input_excel = "Only_Car.xlsx"  # 输入文件名
    # output_excel = "A_Only_Car_0922_1.xlsx"  # 输出文件名
    # input_excel = "Q_Generated_MultiQuestion.xlsx"  # 输入文件名
    # output_excel = "Q_Generated_MultiQuestion_0922_1.xlsx"  # 输出文件名

    process_excel(input_excel, output_excel)
    # 打印结束时间
    end_time = time.time()
    print("开始执行时间：", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(start_time)))
    print("结束执行时间：", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(end_time)))
    print(f"总耗时: {end_time - start_time:.2f} 秒")

