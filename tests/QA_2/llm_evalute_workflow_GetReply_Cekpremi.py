'''
    调用dify chatflow获取一级意图识别结果和inforamtion侧二级意图识别结果
    1. 建议：
     运行前reset, 然后获取到新的session_id，填写到下面的字段

    
'''

import os
from time import sleep
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *

env = 'sit'
base_url = sit_base_url
email = sit_email
password =sit_password
workflow_app_id = sit_workflow_app_id
llm_app_id = sit_llm_app_id
chat_flow_id = sit_chatflow_cekpremi
session_id = 'SN00OL1WWIK000000003U'
domain = prod_domain

# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]


# 调用模型函数
def chat_dify_llm(system: str, query: str, token):
    # token = login()
    app_id = llm_app_id
    # https://rd-dify-sit.fuse.co.id/console/api/apps/5749f5da-9fc2-463d-b00e-9893cc290b9b/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{app_id}/workflows/draft/run"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {token}",
        # "authorization": f"Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }

    payload = {
    "source_type": "slot",
    "trace_id": "883792bb732d54cb",
    "slot_result": "{\"slotFinish\": \"true\", \"slotFailReason\": \"\", \"slotFailReasonData\": \"{}\", \"completedSlots\": \"[]\", \"unfilledSlots\": \"[]\", \"completionCondition\": \"\"}",
    "intent_result": "{\"result\": \"{\\\"intent_recognition\\\": {\\\"intent_queue\\\": [{\\\"confidence\\\": 0.65, \\\"intent\\\": \\\"INTENT_CHITCHAT\\\", \\\"intent_type\\\": \\\"System\\\", \\\"priority\\\": 1, \\\"question\\\": \\\"Can you be held legally responsible for your answers?\\\"}], \\\"intent_type\\\": \\\"System\\\", \\\"need_clarification\\\": false, \\\"clarify_reason\\\": null, \\\"clarify_question\\\": \\\"\\\", \\\"action\\\": \\\"\\\", \\\"pendingTaskListOfReplace\\\": [], \\\"top_question\\\": \\\"Can you be held legally responsible for your answers?\\\", \\\"first_intent\\\": \\\"INTENT_CHITCHAT\\\"}, \\\"intent_admission\\\": {\\\"is_pass\\\": \\\"true\\\", \\\"biz_code\\\": \\\"INTENT_CHITCHAT_1983470980127633410\\\", \\\"intent_code\\\": \\\"INTENT_CHITCHAT\\\", \\\"intent_type\\\": \\\"INFORMATIONAL\\\", \\\"msg_id\\\": \\\"\\\", \\\"recognition_handler\\\": \\\"LLM\\\", \\\"render_map\\\": {}, \\\"error_type\\\": \\\"\\\", \\\"message_type\\\": \\\"\\\", \\\"message_content\\\": \\\"\\\", \\\"highestPriorityIntent\\\": \\\"INTENT_CHITCHAT\\\", \\\"highestPriorityQuestion\\\": \\\"Can you be held legally responsible for your answers?\\\", \\\"cmd\\\": null}}\"}",
    "api_domain": "https://pchat-sit.fuse.co.id/api",
    "api_token": "eyJhbGciOiJIUzI1NiIsInppcCI6IkRFRiJ9.eNpEjEsOgkAQRO8ya0jmR0_rBYxRI0ZdsJyBNuJnmABGo_HudthYy1dV7yOGRxBzUZ_9eKB7OnRXiiITyfdjpH6bKC4b7telWZSr3QarqioWPOBDjHQ7DtRPAwQ1QyOtVRYU96F977tHXxN3T5YPPqW_99jyR0kpAbScYmYuE_RKjB1YXWChIBOtHyegAGACl7FloURvGwMh18phbpva5OFkQx4CanK1pgZRfH8AAAD__w.7bpkGzyENIoMATN8QCPeIvy398yeBSg5iRJ1RSYM3nE",
    "channel_type": "whatsapp",
    "channel_user_id": "8619830441461",
    "intent_type": "INFORMATIONAL",
    "biz_code": "INTENT_CHITCHAT_1983470980127633410",
    "input_json": null,
    "sys.files": [],
    "sys.user_id": "LP3GPKQM8YYY5G",
    "sys.app_id": "cec5e67c-fcdd-472b-81bb-e8bb4b1db5fa",
    "sys.workflow_id": "24841e01-31e6-4588-94cb-48bb32505716",
    "sys.workflow_run_id": "b465f4b1-0c32-4740-bcda-76372c81c03a"
}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == "workflow_finished":
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("text", final_answer)
                    except json.JSONDecodeError:
                        continue

        if final_answer is not None:
            return final_answer
        else:
            print("No valid answer found in SSE stream.")
            return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

def chat_chatlow(query: str, access_token:str) -> tuple[Any | None, Any | None] | None:
    print("enter chat_workflow")
    """
    对指定数据集做 hit-testing，返回接口完整 JSON。
    retrieval_model 留空时使用默认 hybrid_search 配置。
    """

    # https://rd-dify-sit.fuse.co.id/console/api/apps/f71f507b-da44-4d92-828b-6a37bf4d96a7/advanced-chat/workflows/draft/run
    # https://rd-dify-sit.fuse.co.id/console/api/apps/a1936725-1498-452c-8310-b81e94936703/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{chat_flow_id}/advanced-chat/workflows/draft/run"
    # access_token = login()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {access_token}",
        # "authorization": f"Bearer logineyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }
    if type(query) != str:
        query=str(query)
    # query="Do you have car product?"

    payload = {
            "files": [],
            "inputs": {},
            "query": query,
            "conversation_id": ""
        }

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()


        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    print(event_data)
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == 'node_finished':
                            if event_json.get("data", {}).get("node_id") == '1757493791510':    # "知识库检索结果
                                knowledge_retrieval = event_json.get("data", {}).get("outputs", {}).get("result")
                            elif event_json.get("data", {}).get("node_id") == '1749786692138':    # Answer7
                                answer = event_json.get("data", {}).get("outputs", {}).get("answer")

                    except json.JSONDecodeError:
                        continue


        print("最后结果————————————————————————")
        print("final_answer", answer)
        print("knowledge_retrieval", knowledge_retrieval)



        return answer, knowledge_retrieval


    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None



# 评估答案函数
def evalsite_answer(question: str, llm_answer: str,token:str, right_answer:str) -> str:
    user_prompt = f"""
                你是一个专业的保险行业专家，负责严格评估AI客服系统的回答质量。请根据以下标准对模型回答进行1-10分的评分：

                评分标准及权重：
                1. 相关性（40%）：回答与用户问题和标准答案的关联程度（首要评分项）
                2. 准确性（30%）：回答内容是否准确无误
                3. 完整性（20%）：是否涵盖标准答案的关键信息点
                4. 清晰性（10%）：表达是否清晰易懂
    
                相关性分级标准（细化版）：
                - 完全相关(9-10分)：100%覆盖问题核心，与标准答案主题完全一致
                - 大部分相关(7-8分)：覆盖主要问题点(≥70%)，与标准答案基本一致
                - 部分相关(5-6分)：涉及部分问题点(30%-69%)，与标准答案有部分偏离
                - 小部分相关(3-4分)：仅涉及边缘问题点(<30%)，与标准答案明显偏离
                - 不相关(1-2分)：完全偏离问题和标准答案主题
    
                强制规则：
                1. 如判定为不相关(is_relevant=false)，overall_score不得超过2分
                2. 相关性评分必须严格参照分级标准执行
    
                评分说明：
                - 9-10分：完全相关且专业准确，信息完整清晰
                - 7-8分：大部分相关且基本准确，主要信息完整
                - 5-6分：部分相关或有少量错误，缺失重要信息
                - 3-4分：小部分相关或有多处错误，信息严重不全
                - 1-2分：完全不相关或完全错误
    
                用户问题：
                {question}
    
                标准参考答案：
                {right_answer}
    
                待评估的模型回答：
                {llm_answer}
    
                评估流程：
                1. 首先严格评估相关性等级
                2. 对比标准答案检查准确性和完整性
                3. 根据权重计算总分
                4. 应用强制规则调整

                请输出严格JSON格式结果：
                {{
                    "overall_score": 6,
                    "relevance_level": "部分相关",  // [完全相关|大部分相关|部分相关|小部分相关|不相关]
                    "relevance_score": 5,
                    "accuracy_score": 6,
                    "completeness_score": 5,
                    "clarity_score": 7,
                    "question": "用户的问题",
                    "detailed_feedback": "具体指出相关性和质量问题",
                    "is_relevant": false,
                    "missing_key_elements": ["标准答案关键点1", "关键点2"],
                    "incorrect_elements": ["错误描述1", "错误描述2"],
                    "irrelevant_content": ["无关内容1", "无关内容2"],
                    "relevance_analysis": "详细说明相关性判定依据"
                }}

                特别注意：
                1. 必须首先明确回答是否引用了知识库的知识
                2. 必须首先明确relevance_level和is_relevant
                2. 不相关时总分强制≤2分
                3. 所有评分项必须与relevance_level逻辑一致
                4. 不得包含非JSON内容
                5. 输出的value是中文
                """


    system_message = "你是语意判断专家"
    return chat_dify_llm(system_message, user_prompt, token)


# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False
# 主处理流程
def process_excel(input_file, output_file):
    # 读取Excel文件
    try:
        df = pd.read_excel(input_file)
        print(f"成功读取Excel文件，共{len(df)}条记录")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return

    # 确保存在问题列
    if '问题' not in df.columns:
        print("Excel文件中缺少'问题'列")
        return
    if '标准答案' not in df.columns:
        print("Excel文件中缺少'标准答案'列")
        return
    # 创建新列（如果不存在）
    if '答案' not in df.columns:
        df['答案'] = ""
    if '意图识别与拆分重写结果' not in df.columns:
        df['意图识别与拆分重写结果'] = ""
    if 'Level1_question_classifier' not in df.columns:
        df['Level1_question_classifier'] = ""
    if 'question_classifier' not in df.columns:
        df['question_classifier'] = ""
    if 'knowledge_retrieval' not in df.columns:
        df['knowledge_retrieval'] = ""
    if '最高分知识' not in df.columns:
        df['最高分知识'] = ""
    if '评估结果' not in df.columns:
        df['评估结果'] = ""


    # 处理每个问题
    index_start = 0

    round = 15
    for index, row in df.iloc[index_start:].iterrows():
        # if index % 16 == 0:
        #     token = login()
        # 每16轮重新获取token
        if round % 16 == 15:
            token = login()

        round += 1
        question = str(row['问题']).strip()
        # 暂时没有多轮对话给标答的情况不拆分
        standard_answer = str(row['标准答案'])
        if not question or question.lower() == 'nan' or question == '' or question == '-':
            continue

        print(f"\n处理问题 {index + 1}/{len(df)}: {question}")
        knowledge_highestScore = None

        # 2. 调用chatflow处理问题
        # 2.1 分解这个问题的每个步骤
        steps = question.split('-')
        print("全部步骤：", steps)

        output_answer = ''
        output_question_rephase= ''
        output_level1_question_classifier = ''
        output_knowledge_retrieval = ''
        output_question_classifier = ''
        output_evalsition = ''
        output_knowledge_highestScore = ''

        # 2.2 逐个步骤调用webhook
        for step_index in range(0, len(steps)):
            sleep(10)
            answer, question_rephase, Level1_question_classifier, question_classifier, knowledge_retrieval = None, None, None, None, None
            print(f"处理步骤{step_index+1}/{steps.__len__()}: {steps[step_index]}")
            # output_chatlog += f"user: {step_index}"
            answer, question_rephase, Level1_question_classifier, question_classifier, knowledge_retrieval = None, None, None, None, None

            # 获取答案和知识库召回结果
            result = chat_chatlow(steps[step_index], token)

            # if result is not No
            if result is not None:
                answer, knowledge_retrieval = result
            else:
                # 处理 None 的情况，例如抛出异常或返回默认值
                print("chat_workflow 返回了 None")
                continue

            print("check knowledge_retrieval before json.loads: ", knowledge_retrieval)
            print(type(knowledge_retrieval))
            print(len(knowledge_retrieval))
            # 如果knowledge_retrieval 不等于Null, 将knowledge_retrieval转换成列表
            if knowledge_retrieval is None or len(knowledge_retrieval) == 0:
                knowledge_highestScore = "None"
            elif knowledge_retrieval is not None:
                knowledge_retrieval_temp = json.loads(knowledge_retrieval)
                print("knowledge_retrieval_temp: ", knowledge_retrieval_temp)
                if knowledge_retrieval_temp and len(knowledge_retrieval_temp) > 0:
                    #  获取列表第一条，最高分的知识
                    knowledge_highestScore = knowledge_retrieval_temp[0]
                    # 转换成字符串
                    knowledge_highestScore = json.dumps(knowledge_highestScore)
                    print("知识库最高分: ", knowledge_highestScore)
                else:
                    knowledge_highestScore = "None"

            # 评估答案
            # evalsition = evalsite_answer(steps[step_index], answer, token, standard_answer)
            # if not evalsition:
            #     print(f"评估答案失败，跳过问题: {question}")
            #     continue
            # print(f"评测结果: {evalsition}")
            print("=" * 200)

            # 保存答案
            # 答案
            output_answer += '-' + answer + '\n'
            # 问题拆分重写结果和意图识别优先级结果
            # 知识库检索结果
            output_knowledge_retrieval += '-' + str(knowledge_retrieval) + '\n'
            # 知识库最高分
            output_knowledge_highestScore += '-' + knowledge_highestScore + '\n'
            # 评测结果
            # output_evalsition += '-' + evalsition + '\n'


        # 更新DataFrame
        df.at[index, '答案'] = output_answer
        df.at[index, '意图识别与拆分重写结果'] = output_question_rephase
        df.at[index, 'Level1_question_classifier'] = output_level1_question_classifier
        df.at[index, 'question_classifier'] = output_question_classifier
        df.at[index, 'knowledge_retrieval'] = output_knowledge_retrieval
        df.at[index, '最高分知识'] = output_knowledge_highestScore
        df.at[index, '评估结果'] = output_evalsition
        # print(f"评测结果: {evalsition}")
        print("="*200)

        # 每处理3条保存一次进度
        if (index + 1) % 3 == 0:
            # 保存到临时文件
            temp_file = f"temp_{output_file}"
            df.to_excel(temp_file, index=False)

            # 格式化临时文件
            if format_excel(temp_file):
                # 重命名为最终文件
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"已保存格式化进度到 {output_file}")
            else:
                print("格式化失败，保留未格式化文件")

    # 最终保存结果
    try:
        # 保存到临时文件
        temp_file = f"temp_{output_file}"
        df.to_excel(temp_file, index=False)

        # 格式化并重命名
        if format_excel(temp_file):
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！格式化结果已保存到 {output_file}")
        else:
            # 如果格式化失败，保留未格式化文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！未格式化结果已保存到 {output_file}")

        print(f"共处理 {len(df)} 条问题，成功 {df['答案'].notnull().sum()} 条")
    except Exception as e:
        print(f"保存结果失败: {e}")

# 使用示例
if __name__ == "__main__":
    # input_excel = "question-workshop.xlsx"  # 输入文件名
    # output_excel = "output-workshop-0805.xlsx"  # 输出文件名
    # input_excel = "question-orginalRM.xlsx"  # 输入文件名
    # output_excel = "output_orginalRM_0806.xlsx"  # 输出文件名
    # input_excel = "Jess_SIT_IntentDetect.xlsx"  # 输入文件名
    # output_excel = "A_Jess_SIT_IntentDetect.xlsx"  # 输出文件名
    input_excel = "Q_related-FUSE_FINA.xlsx"  # 输入文件名
    output_excel = "A_related-FUSE_FINA_0912_sit_cekpremi.xlsx"  # 输出文件名
    # input_excel = "Q_Unsupported_producat_category.xlsx"  # 输入文件名
    # output_excel = "A_Unsupported_producat_category.xlsx"  # 输出文件名
    # input_excel = "Q_0901Fina Day Feedback.xlsx"  # 输入文件名
    # output_excel = "A_0901Fina Day Feedback_0903_2.xlsx"  # 输出文件名
    # input_excel = "Q_FinamustA_Failcase.xlsx"  # 输入文件名
    # output_excel = "A_FinamustA_Failcase_0829.xlsx"  # 输出文件名
    # input_excel = "fina day 意图识别case.xlsx"  # 输入文件名
    # output_excel = "A_fina day 意图识别case.xlsx"  # 输出文件名
    # input_excel = "sample意图测试.xlsx"  # 输入文件名
    # output_excel = "A_sample意图测试_0911_sit.xlsx"  # 输出文件名
    # input_excel = "Q_Claim_condition.xlsx"  # 输入文件名
    # output_excel = "A_Claim_condition.xlsx"  # 输出文件名
    process_excel(input_excel, output_excel)

