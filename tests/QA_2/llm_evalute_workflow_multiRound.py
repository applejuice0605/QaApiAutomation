import os
from time import sleep
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *
from tests.QA_2.util.llmEvaHandler import evalsite_answer_withStandAns


base_url = sit_base_url
email = sit_email
password =sit_password
workflow_app_id =sit_workflow_app_id
llm_app_id = sit_llm_app_id
webhook_url = uat_webhook_url
wa_id = whatsapp_id

env = 'sit'
base_url = sit_base_url
email = sit_email
password =sit_password
workflow_app_id = sit_workflow_app_id
llm_app_id = sit_llm_app_id
chat_flow_id = sit_sf_v2
session_id = 'SN00OQUB02Z0000000041'
domain = prod_domain
user_info = sit_user_info


# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]

# 调用模型函数
def chat_dify_llm(system: str, query: str, token):
    # token = login()
    app_id = llm_app_id
    # https://rd-dify-sit.fuse.co.id/console/api/apps/5749f5da-9fc2-463d-b00e-9893cc290b9b/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{app_id}/workflows/draft/run"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {token}",
        # "authorization": f"Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }

    payload = {"inputs":{"user":query,"system":system},"files":[]}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == "workflow_finished":
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("text", final_answer)
                    except json.JSONDecodeError:
                        continue

        if final_answer is not None:
            return final_answer
        else:
            print("No valid answer found in SSE stream.")
            return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

def chat_workflow(query: str, access_token:str) -> tuple[Any | None, Any | None, Any | None] | None:
    print("enter chat_workflow")
    """
    对指定数据集做 hit-testing，返回接口完整 JSON。
    retrieval_model 留空时使用默认 hybrid_search 配置。
    """
    # https://rd-dify-sit.fuse.co.id/console/api/apps/a1936725-1498-452c-8310-b81e94936703/workflows/draft/run

    url = f"{base_url.rstrip('/')}/api/apps/{workflow_app_id}/workflows/draft/run"
    # access_token = login()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {access_token}",
        # "authorization": f"Bearer logineyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }
    if type(query) != str:
        query=str(query)
    # query="Do you have car product?"

    payload = {"object":"whatsapp_business_account","entry":[{"id":"1851481102271082","changes":[{"value":{"messaging_product":"whatsapp","metadata":{"display_phone_number":"6285283239812","phone_number_id":"539655932572544"},"contacts":[{"profile":{"name":"nora 2号"},"wa_id":"8619830441461"}],"messages":[{"from":"8619830441461","id":"wamid.HBgNODYxOTgzMDQ0MTQ2MRUCABIYIDdFM0M5QUM4NzAxQ0ZCM0I1QkZFRkYyODNGOEEwNkU5AA==","timestamp":"1751018237","text":{"body": query},"type":"text"}]},"field":"messages"}]}]}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        knowledge_retrieval = None
        question_classifier = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    print(event_data)
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == 'node_finished':
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                            if event_json.get("data", {}).get("node_type") == 'knowledge-retrieval':
                                knowledge_retrieval = event_json.get("data", {}).get("outputs", {}).get("result", [])
                                print(knowledge_retrieval)
                            elif event_json.get("data", {}).get("node_type") == 'question-classifier':
                                question_classifier = event_json.get("data", {}).get("outputs", {}).get("class_name")
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("answer", final_answer)
                    except json.JSONDecodeError:
                        continue
        print(knowledge_retrieval)

        # 使用 json.dumps() 将字典转换为 JSON 格式的字符
        knowledge_retrieval = json.dumps(knowledge_retrieval)

        if final_answer is not None and question_classifier is not None and knowledge_retrieval is not None:
            print("knowledge_retrieval："+knowledge_retrieval)
            return final_answer, question_classifier, knowledge_retrieval
        else:
            print("No valid answer found in SSE stream.")
            return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

# 评估答案函数
def evaluate_answer(question: str, llm_answer: str, right_answer: str, token:str) -> str:
            # 负责严格评估AI客服系统的多轮/单轮对话的每个问题对应的回答质量。
            user_prompt = f"""
                # Role
                你是一名资深的保险专家，你的任务是对另一个AI模型（Gemini Flash）生成的答案进行严格、公正的质量评估。
                
                ## 评估背景
                - **待评估的问题：** “{question}”
                - **待评估的答案（来自被测试模型如Gemini Flash）：** “{llm_answer}”
                - **标准答案（如有，请提供）：** “{right_answer}”
                
                ## 核心指令
                请根据是否提供了“标准答案”来决定评估模式：
                - **【模式一：客观比对模式】**：如果提供了明确的标准答案，请严格比对待评估答案与标准答案的一致性。
                - **【模式二：主观评估模式】**：如果未提供标准答案（标注为‘无’），请基于通用高质量回答标准进行评估。
                
                ## 评估流程与标准
                               
                ### 【模式一：客观比对模式】（当有标准答案时）
                请从以下三个维度评估待评估答案与标准答案的吻合度（总分100分）：
                - **相关性 (30分)**：答案是否直接针对问题，是否答非所问？
                - **准确性 (50分)**：答案中的事实、数据、概念是否与标准答案完全一致？是否存在错误？
                - **完整性 (20分)**：是否涵盖了标准答案中的所有关键要点？是否有重大遗漏？
                
                ### 【模式二：主观评估模式】（当无标准答案时）
                从以下五个核心维度对上述答案进行逐一分析和评分（每个维度满分20分，总分100分）。
                - **a. 事实准确性与专业性 (20分)**
                  - 答案中的保险术语（如免赔额、现金价值、受益人、责任免除等）使用是否准确？
                  - 所述保险知识、原理、法律法规（如《保险法》相关原则）是否正确？
                  - 是否存在事实性错误或过时信息？
                - **b. 逻辑严谨性与完整性 (20分)**
                  - 推理过程是否清晰、有逻辑？能否清晰地解释“为什么”？
                  - 是否涵盖了问题中所有关键点？是否考虑了不同的情景或例外情况（如免责条款）？
                  - 答案结构是否完整（如：先解释概念，再分析案例，最后总结）？
                - **c. 清晰度与可读性 (20分)**
                  - 答案是否组织良好、易于理解？是否使用了分点、分段等格式？
                  - 能否用通俗的语言向非专业人士解释复杂概念？
                  - 语言是否流畅、简洁，没有不必要的赘述？
                - **d. 谨慎性与负责任程度 (20分)**
                  - 答案是否包含了必要的免责声明（如“具体请以保险合同条款为准”、“建议咨询专业的保险顾问或保险公司”）？
                  - 是否避免了给出绝对的、可能构成医疗或法律建议的结论？
                  - 态度是否中立、客观，不会误导用户？
                - **e. 相关性与实用性 (20分)**
                  - 答案是否直接回答了核心问题，没有答非所问或避重就轻？
                  - 对于用户提出的问题，给出的信息是否具有实际操作价值？能否真正帮助用户解决疑惑或做出下一步决策？
                                  
                ## 输出格式要求
                请严格按照以下格式组织你的最终评估报告：
                
                **【评估模式】**
                [在此声明采用的评估模式：模式一（客观比对）或 模式二（主观评估）]
                
                **【总体评分】**
                [总分] / 100
                
                **【分维度评分】**
                - [维度1名称]: [得分] / [满分]
                - [维度2名称]: [得分] / [满分]
                - [维度3名称]: [得分] / [满分]
                - [维度4名称（如有）]: [得分] / [满分]
                
                **【详细分析】**
                - **优点：** 列出待评估答案中表现突出的部分。
                - **缺点与错误：** 列出待评估答案中存在的具体问题、错误或遗漏。在模式一下，必须明确指出与标准答案不符之处。
                - **改进建议：** 提供1-2句具体的优化建议。
                
                # 特别注意：
                # 1. 用中文输出评估结果
                """
            system_message = "你是语意判断专家"
            return chat_dify_llm(system_message, user_prompt, token)


# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False

# 主处理流程
def process_excel(input_file, output_file):

    resetSession()
    # 读取Excel文件
    try
        df = pd.read_excel(input_file)
        print(f"成功读取Excel文件，共{len(df)}条记录")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return

    # 确保存在问题列
    if '问题' not in df.columns:
        print("Excel文件中缺少'问题'列")
        return

    # 创建新列（如果不存在）
    if '答案' not in df.columns:
        df['答案'] = ""
    if '评估结果' not in df.columns:
        df['评估结果'] = ""

    # 处理每个问题
    index_start = 0

    round = 15
    for index, row in df.iloc[index_start:].iterrows():
        # 1. 重置Session
        # resetSession()

        # if index % 16 == 0:
        #     token = login()
        # 每16轮重新获取token
        if round % 16 == 15:
            token = login()

        round += 1
        question = str(row['问题']).strip()
        standard = str(row['标准答案']).strip()
        # standard = ''
        if not question or question.lower() == 'nan' or question == '' or question == '-':
            continue

        print(f"\n处理问题 {index + 1}/{len(df)}: {question}")

        # 2. 调用webhook处理问题
        # 2.1 分解这个问题的每个步骤
        steps = question.split('- ')
        print("全部步骤：", steps)

        output_answer = ''
        output_knowledge_retrieval = ''
        output_question_classifier = ''
        output_evaluation = ''
        output_chatlog = ''

        # 2.2 逐个步骤调用webhook
        # for step_index in range(1, len(steps)):
        for step_index in range(0, len(steps)):
            print(f"处理步骤{step_index}/{steps.__len__()-1}: {steps[step_index]}")
            output_chatlog += f"user: {step_index}"
            trace_id = None
            answer = None
            step = steps[step_index].strip()

            # 2.3 调用webhook
            if step == '{send GPS}':
                msg_body = {
                                "from": "8619830441461",
                                "id": "wamid.HBgNODYxOTgzMDQ0MTQ2MRUCABIYIDkyNkFBMDg4MzcyRDI2QjNCNTkzQTRCMjE3RjgxODFBAA==",
                                "timestamp": "1754601252",
                                "location": {
                                    "latitude": 22.5250867,
                                    "longitude": 113.9215657
                                },
                                "type": "location"
                            }
            else:
                msg_body = {
                                "from": "8619830441461",
                                "id": "wamid.HBgNODYxOTgzMDQ0MTQ2MRUCABIYFjNFQjA4RTgwNEMwQkRGNDlFMkZEQTAA",
                                "timestamp": "1754530501",
                                "text": {
                                    "body": step
                                },
                                "type": "text"
                            }

            trace_id = request_webhook(msg_body)

            print("trace_id:", trace_id)
            # 2.4 请求数据库，查询聊天记录
            answer = '在数据库中没有找到数据'
            print(answer.find('在数据库中没有找到数据') != -1)
            while answer.find('在数据库中没有找到数据') != -1:
                answer = getAnswerFromDB_bytraceId(trace_id)
                print("answer:", answer)
                if answer is not None:
                    print(answer.find('在数据库中没有找到数据') != -1)
                    sleep(2)
                else:
                    answer = '在数据库中没有找到数据'
            # if trace_id is not None:
            #     print(step_index)
            #     # 休眠5s
            #     sleep(20)
            #     for i in range(1,5):
            #         answer = getAnswerFromDB_bytraceId(trace_id)
            #         # 如果answer 不以"在数据库中没有找到数据"结尾，跳出循环
            #         if answer.find('在数据库中没有找到数据') == -1:
            #             break
            #         else:
            #             i += 1
            # else:
            #     # 处理 None 的情况，例如抛出异常或返回默认值
            #     answer = str(step_index) + '. ' + trace_id + '\n'
            #     # raise ValueError("webhook 返回了 None")

            print(f"工作流的回答:\n {answer}")
            output_chatlog += f"Bot: {answer}\n"
            # print(f"评测结果: {evaluation}")
            print("=" * 200)

            # 保存答案
            if answer is None:
                raise ValueError("数据库没有找到答案")

            output_answer += answer
            # output_knowledge_retrieval = '- ' + knowledge_retrieval + '\n'
            # output_question_classifier = '- ' + question_classifier + '\n'
            # output_evaluation = '- ' + evaluation + '\n'

        # 评估答案
        # TODO：重写机器人评分的标准和机制:
        #     看意图分类对不对
        #     看是否有用知识库的内容回答问题
        #     是的话，给出引用的文件和内容
        #     再用通用标准评分
        # evaluation = evalsite_answer_withStandAns(base_url=base_url, app_id=llm_app_id, token=token, question=step, llm_answer=answer, right_answer=standard)
        # if not evaluation:
        #     print(f"评估答案失败，跳过问题: {question}")
        #     continue
        # print(f"评测结果: {evaluation}")
        # 更新DataFrame
        print("更新dataframe")
        print("output_answer:{output_answer}", output_answer)
        # print("evaluation:{evaluation}", evaluation)
        df.at[index, '答案'] = output_answer
        # df.at[index, 'question_classifier'] = output_knowledge_retrieval
        # df.at[index, 'knowledge_retrieval'] = output_question_classifier
        # df.at[index, '评估结果'] = evaluation


        # 每处理3条保存一次进度
        if (index + 1) % 3 == 0:
            # 保存到临时文件
            temp_file = f"temp_{output_file}"
            df.to_excel(temp_file, index=False)

            # 格式化临时文件
            if format_excel(temp_file):
                # 重命名为最终文件
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"已保存格式化进度到 {output_file}")
            else:
                print("格式化失败，保留未格式化文件")

    # 最终保存结果
    try:
        # 保存到临时文件
        temp_file = f"temp_{output_file}"
        df.to_excel(temp_file, index=False)

        # 格式化并重命名
        if format_excel(temp_file):
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！格式化结果已保存到 {output_file}")
        else:
            # 如果格式化失败，保留未格式化文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！未格式化结果已保存到 {output_file}")

        print(f"共处理 {len(df)} 条问题，成功 {df['答案'].notnull().sum()} 条")
    except Exception as e:
        print(f"保存结果失败: {e}")


def resetSession():
    # 调用webhook发送“resetSession”指令
    request_webhook("restart")

    # request_webhook("1234")


def request_webhook(msg_body):
    print("start request_webhook")
    """
    调用webhook发送文本消息
    """
    # https://pchat-uat.fuse.co.id/api/ai/chatbot/whatsapp/webhook

    url = f"{webhook_url.rstrip('/')}/api/ai/chatbot/whatsapp/webhook"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }
    # if type(msg_body) != str:
    #     query = str(msg_body)
    # user_input="reset session"

    # TODO：动态获取messgae from的wa_id, wa_token，business_account_id
    payload = {
            "object": "whatsapp_business_account",
            "entry": [
                {
                    "id": "1851481102271082",
                    "changes": [
                        {
                            "value": {
                                "messaging_product": "whatsapp",
                                "metadata": {
                                    "display_phone_number": "6285283239812",
                                    "phone_number_id": "539655932572544"
                                },
                                "contacts": [
                                    {
                                        "profile": {
                                            "name": "nora 2号"
                                        },
                                        "wa_id": "8613434915136"
                                    }
                                ],
                                "messages": [
                                    msg_body
                                ]
                            },
                            "field": "messages"
                        }
                    ]
                }
            ]
        }

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        # print(resp)
        # resp.raise_for_status()
        # print(resp.text)

    except requests.RequestException as e:
        print(f"Error calling webhook: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

    return resp.text


def getAnswerFromDB_bytraceId(trace_id):
    print("start queryDB_bytraceId")

    """
    调用webhook发送文本消息
    """
    # https://pchat-uat.fuse.co.id/api/ai/chatbot/whatsapp/webhook

    url = f"https://rd-dms.fuseinsurtech.com/query/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/x-www-form-urlencoded",
        "cookie": "__adroll_fpc=6be6a387174f9054dd65546054f7611d-1760584004092; _fbp=fb.1.1760584006506.534717559103612633; csrftoken=VoKNJytm76En0OquInL7zBqtvZIHjKhv7uRgSwMEDuuZs9EJRhpY5CxsvIrSwXAL; sessionid=759b25ydip8mntf00xn0bc0loi9lty12",
        "x-csrftoken": "VoKNJytm76En0OquInL7zBqtvZIHjKhv7uRgSwMEDuuZs9EJRhpY5CxsvIrSwXAL"
    }
    if type(trace_id) != str:
        trace_id = str(trace_id)

    # trace_id = '8c6f64a18338a8da'

    sql_content = f"select reply from message.whatsapp_chat_record where 1=1 and trace_id = '{trace_id}' order by uid desc limit 1"

    # TODO：动态获取messgae from的wa_id, wa_token，business_account_id
    payload = {
        "instance_name" : "ID_UAT_CORE_MYSQL8.0",
        "db_name" : "message",
        "schema_name" : None,
        "tb_name": None,
        "sql_content" : sql_content,
        "limit_num" : "100"
    }
    try:
        resp = requests.post(url, headers=headers, data=payload)
        resp = requests.post(url, headers=headers, data=payload)
        print(resp.json())

    except requests.RequestException as e:
        print(f"Error calling db: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return "Error calling db"

    if resp.json()['data']['affected_rows'] == 0:
        answer = trace_id + "在数据库中没有找到数据"
    elif len(resp.json()['data']['rows'][0]) == 0:
        answer = trace_id + "在数据库中没有找到数据"
    else:
        answer = resp.json()['data']['rows'][0][0]

    print(answer)
    return answer

# 使用示例
if __name__ == "__main__":
    # input_excel = "question_workshop_multiround.xlsx"  # 输入文件名
    # output_excel = "output_workshop_multiround-0805.xlsx"  # 输出文件名


    # input_excel = "question_FQA_multi-round.xlsx"  # 输入文件名
    # output_excel = "output_FQA_multi-round.xlsx"  # 输出文件名
    # input_excel = "question_EN.xlsx"  # 输入文件名
    # output_excel = "output_EN.xlsx"  # 输出文件名
    input_excel = "ID_SF_BatchTest_webhook_1105.xlsx"  # 输入文件名
    output_excel = "A_ID_SF_BatchTest_webhook_1107_01.xlsx.xlsx"  # 输出文件名
    # input_excel = "Q_温度.xlsx"  # 输入文件名
    # output_excel = "A_温度_0921_3.xlsx"  # 输出文件名
    process_excel(input_excel, output_excel)

