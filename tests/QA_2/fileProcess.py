# 定义一个文件处理类，包括功能：读取文件、写文件、格式化文件

import os
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *


class FileProcess:
    def __init__(self, file_path):
        self.file_path = file_path
        self.file = open(file_path, 'r+')
        self.lines = self.file.readlines()
        self.file.close()
    def read_file(self):
        return self.lines
    def write_file(self, content):
        self.file = open(self.file_path, 'w')
        self.file.write(content)
        self.file.close()

    @staticmethod
    def format_excel(file_path):
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            ws = wb.active

            # 设置所有单元格自动换行和顶部对齐
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # 自动调整列宽（限制最大宽度）
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        value = str(cell.value) if cell.value else ""
                        # 计算单元格内容的长度（考虑换行）
                        lines = value.split('\n')
                        max_line_length = max(len(line) for line in lines) if lines else 0
                        if max_line_length > max_length:
                            max_length = max_line_length
                    except:
                        pass

                # 设置合理的列宽（50-100个字符）
                adjusted_width = min(max_length + 2, 100)
                adjusted_width = max(adjusted_width, 15)  # 最小宽度15
                ws.column_dimensions[column].width = adjusted_width

            # 设置特定列的宽度
            for col_letter in ['A', 'B', 'C', 'D']:
                if col_letter in ws.column_dimensions:
                    if col_letter == 'A':  # 问题列
                        ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                    elif col_letter == 'B':  # 答案列
                        ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                    elif col_letter == 'C':  # 评估结果列
                        # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                        ws.column_dimensions[col_letter].width = 200

            # 自动调整行高（基于内容行数）
            for row in ws.iter_rows():
                max_lines = 1
                for cell in row:
                    if cell.value:
                        # 计算单元格中的行数
                        lines = str(cell.value).split('\n')
                        if len(lines) > max_lines:
                            max_lines = len(lines)
                # 设置行高（每行15点，最小20点）
                ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

            # 保存格式化后的文件
            wb.save(file_path)
            print(f"已格式化Excel文件: {file_path}")
            return True
        except Exception as e:
            print(f"格式化Excel文件时出错: {e}")
            return False

    @staticmethod
    def real_excel(input_file):
        # 读取Excel文件
        try:
            df = pd.read_excel(input_file)
            print(f"成功读取Excel文件，共{len(df)}条记录")
        except Exception as e:
            print(f"读取Excel文件失败: {e}")
            return

        # 确保存在问题列
        if '问题' not in df.columns:
            print("Excel文件中缺少'问题'列")
            return
        if '标准答案' not in df.columns:
            print("Excel文件中缺少'标准答案'列")
            return
        # 创建新列（如果不存在）
        if '答案' not in df.columns:
            df['答案'] = ""
        # if '意图识别与拆分重写结果' not in df.columns:
        #     df['意图识别与拆分重写结果'] = ""
        # if 'Level1_question_classifier' not in df.columns:
        #     df['Level1_question_classifier'] = ""
        # if 'question_classifier' not in df.columns:
        #     df['question_classifier'] = ""
        # if 'knowledge_retrieval' not in df.columns:
        #     df['knowledge_retrieval'] = ""
        # if '最高分知识' not in df.columns:
        #     df['最高分知识'] = ""
        # if '评估结果' not in df.columns:
        #     df['评估结果'] = ""

        return df

    @staticmethod
    def save_result_temp(index, df, output_file):
        try:
            # 每处理3条保存一次进度
            if (index + 1) % 3 == 0:
                # 使用不同的临时文件名
                temp_file = os.path.splitext(output_file)[0] + "_temp" + os.path.splitext(output_file)[1]

                # 确保临时文件不存在
                if os.path.exists(temp_file):
                    os.remove(temp_file)

                # 保存到临时文件
                df.to_excel(temp_file, index=False)
                print(f"已保存临时文件到: {temp_file}")

                # 格式化临时文件
                if FileProcess.format_excel(temp_file):
                    # 确保最终文件不存在
                    if os.path.exists(output_file):
                        os.remove(output_file)

                    # 重命名为最终文件
                    os.rename(temp_file, output_file)
                    print(f"已保存格式化进度到 {output_file}")
                else:
                    print("格式化失败，保留未格式化文件")
                    # 如果格式化失败，删除临时文件
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
        except Exception as e:
            print(f"保存进度时发生错误: {str(e)}")
            # 确保清理临时文件
            if os.path.exists(temp_file):
                os.remove(temp_file)
            raise

    @staticmethod
    def write_to_excel(df, output_file):
        # 最终保存结果
        try:
            # 保存到临时文件
            temp_file = f"temp_{output_file}"
            df.to_excel(temp_file, index=False)

            # 格式化并重命名
            if FileProcess.format_excel(temp_file):
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"\n处理完成！格式化结果已保存到 {output_file}")
            else:
                # 如果格式化失败，保留未格式化文件
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"\n处理完成！未格式化结果已保存到 {output_file}")

            print(f"共处理 {len(df)} 条问题，成功 {df['答案'].notnull().sum()} 条")
        except Exception as e:
            print(f"保存结果失败: {e}")

