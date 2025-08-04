import os
from typing import Dict, Any, Optional, Tuple
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config import *


base_url = uat_base_url
email = uat_email
password =uat_password
workflow_app_id =uat_workflow_app_id
llm_app_id = uat_llm_app_id

# 登录函数获取token
def login() -> str:
    """
    登录 Dify Console，返回 access_token。
    extra_login_fields 可额外传 remember_me/language 等字段。
    """
    url = f"{base_url.rstrip('/')}/api/login"
    payload = {
        "email": email,
        "password": password,
        "language": "zh-Hans",
        "remember_me": True
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "content-type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json()["data"]["access_token"]

# 调用模型函数
def chat_dify_llm(system: str, query: str, token):
    # token = login()
    app_id = llm_app_id
    # https://rd-dify-sit.fuse.co.id/console/api/apps/5749f5da-9fc2-463d-b00e-9893cc290b9b/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{app_id}/workflows/draft/run"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {token}",
        # "authorization": f"Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }

    payload = {"inputs":{"user":query,"system":system},"files":[]}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == "workflow_finished":
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("text", final_answer)
                    except json.JSONDecodeError:
                        continue

        if final_answer is not None:
            return final_answer
        else:
            print("No valid answer found in SSE stream.")
            return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

def chat_workflow(query: str, access_token:str) -> tuple[Any | None, Any | None, Any | None] | None:
    print("enter chat_workflow")
    """
    对指定数据集做 hit-testing，返回接口完整 JSON。
    retrieval_model 留空时使用默认 hybrid_search 配置。
    """
    # https://rd-dify-sit.fuse.co.id/console/api/apps/a1936725-1498-452c-8310-b81e94936703/workflows/draft/run
    url = f"{base_url.rstrip('/')}/api/apps/{workflow_app_id}/workflows/draft/run"
    # access_token = login()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/137.0.0.0 Safari/537.36",
        "authorization": f"Bearer {access_token}",
        # "authorization": f"Bearer logineyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoiNTUyYmRiODktNTlkOC00NDExLWJiZmYtMjM3MGQwOWQ4MTljIiwiZXhwIjoxNzUzNzEwNDI0LCJpc3MiOiJTRUxGX0hPU1RFRCIsInN1YiI6IkNvbnNvbGUgQVBJIFBhc3Nwb3J0In0.uZAtxK0YEYiPyymg_g6ZDp7ckeRPRCbIw00TTj7rSD8",
        "content-type": "application/json",
    }
    if type(query) != str:
        query=str(query)
    # query="Do you have car product?"

    payload = {"inputs": {"question": query}, "files": []}

    try:
        resp = requests.post(url, headers=headers, json=payload, stream=True)
        resp.raise_for_status()

        final_answer = None
        knowledge_retrieval = None
        question_classifier = None
        for line in resp.iter_lines():
            if line:
                decoded_line = line.decode('utf-8').strip()
                if decoded_line.startswith("data: "):
                    event_data = decoded_line[6:]  # 去掉 "data: "
                    print(event_data)
                    try:
                        event_json = json.loads(event_data)
                        if event_json.get("event") == 'node_finished':
                            final_answer = event_json.get("data", {}).get("outputs", {}).get("text")
                            if event_json.get("data", {}).get("node_type") == 'knowledge-retrieval':
                                knowledge_retrieval = event_json.get("data", {}).get("outputs", {}).get("result", [])
                                print(knowledge_retrieval)
                            elif event_json.get("data", {}).get("node_type") == 'question-classifier':
                                question_classifier = event_json.get("data", {}).get("outputs", {}).get("class_name")
                        elif event_json.get("event") == "message_end":
                            final_answer = event_json.get("metadata", {}).get("answer", final_answer)
                    except json.JSONDecodeError:
                        continue
        print(knowledge_retrieval)

        # 使用 json.dumps() 将字典转换为 JSON 格式的字符
        knowledge_retrieval = json.dumps(knowledge_retrieval)

        if final_answer is not None and question_classifier is not None and knowledge_retrieval is not None:
            print("knowledge_retrieval："+knowledge_retrieval)
            return final_answer, question_classifier, knowledge_retrieval
        else:
            print("No valid answer found in SSE stream.")
            return None

    except requests.RequestException as e:
        print(f"Error calling chat: {e}")
        if hasattr(e, 'response') and e.response:
            print(f"Response status: {e.response.status_code}")
            print(f"Response body: {e.response.text}")
        return None

# 评估答案函数
def evaluate_answer(question: str, answer: str,token:str) -> str:
            user_prompt = f"""
                你是一个专业的保险行业专家，负责严格评估AI客服系统的回答质量。请根据以下标准对模型回答进行1-10分的评分：

                评分标准及权重：
                1. 相关性（40%）：回答与用户问题和你的认识关联程度（首要评分项）
                2. 准确性（30%）：回答内容是否准确无误和语言一致性
                3. 完整性（20%）：是否涵盖标你的认识的关键信息点
                4. 清晰性（10%）：表达是否清晰易懂

                相关性分级标准（细化版）：
                - 完全相关(9-10分)：100%覆盖问题核心，与你的认识完全一致，且语言一致
                - 大部分相关(7-8分)：覆盖主要问题点(≥70%)，与你的认识基本一致，且语言一致
                - 部分相关(5-6分)：涉及部分问题点(30%-69%)，与你的认识有部分偏离，且语言一致
                - 小部分相关(3-4分)：仅涉及边缘问题点(<30%)，与你的认识明显偏离
                - 不相关(1-2分)：完全偏离问题和你认识主题

                强制规则：
                1. 如判定为不相关(is_relevant=false)，overall_score不得超过2分
                2. 相关性评分必须严格参照分级标准执行

                评分说明：
                - 9-10分：完全相关且专业准确，信息完整清晰，且语言一致
                - 7-8分：大部分相关且基本准确，主要信息完整，且语言一致
                - 5-6分：部分相关或有少量错误，缺失重要信息
                - 3-4分：小部分相关或有多处错误，信息严重不全
                - 1-2分：完全不相关或完全错误

                用户问题：
                {question}

                待评估的模型回答：
                {answer}

                评估流程：
                1. 首先严格评估相关性等级
                2. 对比你的认识检查准确性和完整性
                3. 根据权重计算总分
                4. 应用强制规则调整

                请输出严格JSON格式结果：
                {{
                    "overall_score": 6,
                    "relevance_level": "部分相关",  // [完全相关|大部分相关|部分相关|小部分相关|不相关]
                    "relevance_score": 5,
                    "accuracy_score": 6,
                    "completeness_score": 5,
                    "clarity_score": 7,
                    "question": "用户的问题",
                    "detailed_feedback": "具体指出相关性和质量问题",
                    "is_relevant": false,
                    "missing_key_elements": ["标准答案关键点1", "关键点2"],
                    "incorrect_elements": ["错误描述1", "错误描述2"],
                    "irrelevant_content": ["无关内容1", "无关内容2"],
                    "relevance_analysis": "详细说明相关性判定依据"
                }}

                特别注意：
                1. 必须首先明确relevance_level和is_relevant
                2. 不相关时总分强制≤2分
                3. 所有评分项必须与relevance_level逻辑一致
                4. 不得包含非JSON内容
                5. 输出的value是中文
                """
            system_message = "你是语意判断专家"
            return chat_dify_llm(system_message, user_prompt, token)


# 格式化Excel文件
def format_excel(file_path):
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置所有单元格自动换行和顶部对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 自动调整列宽（限制最大宽度）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    # 计算单元格内容的长度（考虑换行）
                    lines = value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    if max_line_length > max_length:
                        max_length = max_line_length
                except:
                    pass

            # 设置合理的列宽（50-100个字符）
            adjusted_width = min(max_length + 2, 100)
            adjusted_width = max(adjusted_width, 15)  # 最小宽度15
            ws.column_dimensions[column].width = adjusted_width

        # 设置特定列的宽度
        for col_letter in ['A', 'B', 'C', 'D']:
            if col_letter in ws.column_dimensions:
                if col_letter == 'A':  # 问题列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'B':  # 答案列
                    ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 40)
                elif col_letter == 'C':  # 评估结果列
                    # ws.column_dimensions[col_letter].width = min(ws.column_dimensions[col_letter].width, 200)
                    ws.column_dimensions[col_letter].width = 200

        # 自动调整行高（基于内容行数）
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value:
                    # 计算单元格中的行数
                    lines = str(cell.value).split('\n')
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            # 设置行高（每行15点，最小20点）
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        # 保存格式化后的文件
        wb.save(file_path)
        print(f"已格式化Excel文件: {file_path}")
        return True
    except Exception as e:
        print(f"格式化Excel文件时出错: {e}")
        return False
# 主处理流程
def process_excel(input_file, output_file):
    # 读取Excel文件
    try:
        df = pd.read_excel(input_file)
        print(f"成功读取Excel文件，共{len(df)}条记录")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return

    # 确保存在问题列
    if '问题' not in df.columns:
        print("Excel文件中缺少'问题'列")
        return

    # 创建新列（如果不存在）
    if '答案' not in df.columns:
        df['答案'] = ""
    if '评估结果' not in df.columns:
        df['评估结果'] = ""

    # 处理每个问题
    index_start = 218

    round = 15
    for index, row in df.iloc[index_start:].iterrows():
        # if index % 16 == 0:
        #     token = login()
        # 每16轮重新获取token
        if round % 16 == 15:
            token = login()

        round += 1
        question = str(row['问题']).strip()
        if not question or question.lower() == 'nan':
            continue

        print(f"\n处理问题 {index + 1}/{len(df)}: {question}")

        # 获取答案和知识库召回结果
        result = chat_workflow( question, token)
        if result is not None:
            answer, question_classifier, knowledge_retrieval = result
        else:
            # 处理 None 的情况，例如抛出异常或返回默认值
            raise ValueError("chat_workflow 返回了 None")

        if not answer:
            print(f"获取答案失败，跳过问题: {question}")
            continue

        # 评估答案
        evaluation = evaluate_answer(question, answer, token)
        if not evaluation:
            print(f"评估答案失败，跳过问题: {question}")
            continue

        # 更新DataFrame
        df.at[index, '答案'] = answer
        df.at[index, 'question_classifier'] = question_classifier
        df.at[index, 'knowledge_retrieval'] = knowledge_retrieval
        df.at[index, '评估结果'] = evaluation
        print(f"工作流的回答: {answer}")
        print(f"评测结果: {evaluation}")
        print("="*200)

        # 每处理3条保存一次进度
        if (index + 1) % 3 == 0:
            # 保存到临时文件
            temp_file = f"temp_{output_file}"
            df.to_excel(temp_file, index=False)

            # 格式化临时文件
            if format_excel(temp_file):
                # 重命名为最终文件
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                print(f"已保存格式化进度到 {output_file}")
            else:
                print("格式化失败，保留未格式化文件")

    # 最终保存结果
    try:
        # 保存到临时文件
        temp_file = f"temp_{output_file}"
        df.to_excel(temp_file, index=False)

        # 格式化并重命名
        if format_excel(temp_file):
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！格式化结果已保存到 {output_file}")
        else:
            # 如果格式化失败，保留未格式化文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_file, output_file)
            print(f"\n处理完成！未格式化结果已保存到 {output_file}")

        print(f"共处理 {len(df)} 条问题，成功 {df['答案'].notnull().sum()} 条")
    except Exception as e:
        print(f"保存结果失败: {e}")

# 使用示例
if __name__ == "__main__":
    # input_excel = "output-orginalRM.xlsx"  # 输入文件名
    # output_excel = "output-orginalRM-0731.xlsx"  # 输出文件名
    input_excel = "question.xlsx"  # 输入文件名
    output_excel = "output_0804.xlsx"  # 输出文件名
    process_excel(input_excel, output_excel)

