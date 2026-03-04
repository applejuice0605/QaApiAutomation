import unittest
from unittest.mock import patch, mock_open
from io import StringIO
import json
from openpyxl import Workbook
from ExcelHandler import ExcelHandler
# import ExcelHandler  # 替换为你的模块名
import os

class TestExcelHandler(unittest.TestCase):

    def setUp(self):
        self.excel_handler = ExcelHandler()
        self.test_file_path = 'chatFlow.xlsx'
        self.test_sheet_name = 'Sheet1'
        self.test_data = [
            {"input": "data1", "expected": "data2"},
            {"input": "data3", "expected": "data4"}
        ]

        # Create a test workbook and sheet
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.sheet.title = self.test_sheet_name
        for i, data in enumerate(self.test_data):
            self.sheet.cell(row=i+1, column=1).value = json.dumps(data["input"])
            self.sheet.cell(row=i+1, column=2).value = json.dumps(data["expected"])
        self.wb.save(self.test_file_path)

    def tearDown(self):
        self.wb.close()
        if self.excel_handler.wb:
            self.excel_handler.wb.close()
        try:
            os.remove(self.test_file_path)
        except OSError:
            pass

    @patch('builtins.open', new_callable=mock_open)
    def test_hello(self, mock_open):
        result = self.excel_handler.hello()
        self.assertEqual(result, "Hello, World!")

    @patch('openpyxl.load_workbook')
    def test_open_excel(self, mock_load_workbook):
        mock_load_workbook.return_value = self.wb
        self.excel_handler.open_excel(self.test_file_path, self.test_sheet_name)
        self.assertEqual(self.excel_handler.file_path, self.test_file_path)
        self.assertEqual(self.excel_handler.sheet.title, self.sheet.title)

    def test_get_row_count(self):
        self.excel_handler.open_excel(self.test_file_path, self.test_sheet_name)
        result = self.excel_handler.get_row_count()
        self.assertEqual(result, len(self.test_data))

    def test_read_row_data(self):
        self.excel_handler.open_excel(self.test_file_path, self.test_sheet_name)
        for i, data in enumerate(self.test_data):
            input_data, expected_data = self.excel_handler.read_row_data(i+1)
            self.assertEqual(input_data, data["input"])
            self.assertEqual(expected_data, data["expected"])

    @patch('builtins.open', new_callable=mock_open)
    def test_write_results(self, mock_open):
        self.excel_handler.open_excel(self.test_file_path, self.test_sheet_name)
        for i, data in enumerate(self.test_data):
            output = {"result": "success"}
            actual = True
            self.excel_handler.write_results(i+1, output, actual)
            self.assertEqual(self.sheet.cell(row=i+1, column=3).value, json.dumps(output))
            self.assertEqual(self.sheet.cell(row=i+1, column=4).value, 'pass')

    def test_close_excel(self):
        self.excel_handler.open_excel(self.test_file_path, self.test_sheet_name)
        self.excel_handler.close_excel()
        self.assertIsNone(self.excel_handler.wb)

if __name__ == '__main__':
    unittest.main()
