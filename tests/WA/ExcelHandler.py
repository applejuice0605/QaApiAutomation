import openpyxl
import json
from openpyxl import load_workbook
from robot.api.deco import keyword

class ExcelHandler:
    def __init__(self):
        self.wb = None
        self.sheet = None
        # self.file_path = file_path

    @keyword
    def open_excel(self, file_path, sheet_name):
        # ExcelHandler.__init__(self, file_path)
        self.file_path = file_path
        # 以可写方式打开
        self.wb = load_workbook(file_path, read_only=False)
        self.sheet = self.wb[sheet_name]
        return self.wb

    @keyword
    def get_row_count(self):
        return self.sheet.max_row

    @keyword
    def read_row_data(self, row):
        input_value = str(self.sheet.cell(row=row, column=3).value or "").strip()
        expected_value = str(self.sheet.cell(row=row, column=4).value or "").strip()
        print(input_value)
        print(expected_value)
        # return input_str, expected_str


        # input_value = self.sheet.cell(row=row, column=3).value
        # expected_value = self.sheet.cell(row=row, column=4).value
        # output_value = self.sheet.cell(row=row, column=3).value
        # try:
        #     input_data = json.loads(input_value) if input_value else {}
        # except json.JSONDecodeError:
        #     raise ValueError(f"Input数据格式错误，行{row}列1的内容不是有效的JSON: {input_value}")
        # try:
        #     expected_data = json.loads(expected_value) if expected_value else {}
        # except json.JSONDecodeError:
        #     raise ValueError(f"Expected数据格式错误，行{row}列2的内容不是有效的JSON: {expected_value}")
        return input_value, expected_value

    #获取指定行列的数据
    @keyword
    def get_cell_value(self, row, column):
        row = int(row)
        column = int(column)
        return self.sheet.cell(row=row, column=column).value

    # 在指定行列写入数据
    @keyword
    def write_cell_value(self, row, column, value):
        row = int(row)
        column = int(column)
        self.sheet.cell(row=row, column=column).value = value
        self.wb.save(self.file_path)

    @keyword
    def write_results(self, row, output, actual):
        self.sheet.cell(row=row, column=3).value = json.dumps(output)
        self.sheet.cell(row=row, column=4).value = 'pass' if actual else 'fail'
        self.wb.save(self.file_path)

    @keyword
    def close_excel(self):
        # self.wb.close()
        self.wb = None

